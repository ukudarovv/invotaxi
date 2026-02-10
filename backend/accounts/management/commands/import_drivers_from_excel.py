"""
Management command для импорта водителей из Excel файла
"""
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.conf import settings
from accounts.models import Driver, DriverStatus
from regions.models import Region
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
import os

logger = logging.getLogger(__name__)
User = get_user_model()


class Command(BaseCommand):
    help = '''
    Импортирует водителей из Excel файла.
    
    Поддерживаемые поля (соответствуют форме "Добавить водителя"):
    Обязательные:
    - Имя* (name, имя, фио)
    - Телефон* (phone, телефон)
    - Пароль* (password, пароль) - минимум 8 символов
    - Регион* (region_id, region_title, регион, id региона, название региона)
    - Машина* (car_model, модель машины, машина, автомобиль)
    - Гос. номер* (plate_number, гос номер, номер машины)
    - Вместимость* (capacity, вместимость, мест)
    
    Опциональные:
    - Email (email, почта)
    - Водитель онлайн (is_online, онлайн, online) - по умолчанию False
    - Рейтинг (rating, рейтинг) - по умолчанию 5.0
    
    Пример использования:
    python manage.py import_drivers_from_excel "data/водитель (1).xlsx" --skip-errors
    
    Для генерации шаблона:
    python manage.py import_drivers_from_excel --generate-template
    python manage.py import_drivers_from_excel --generate-template "data/шаблон_водители.xlsx"
    '''

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            nargs='?',
            default='data/водитель (1).xlsx',
            help='Путь к Excel файлу (по умолчанию: data/водитель (1).xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Только валидация, без создания водителей',
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Пропускать строки с ошибками и продолжать',
        )
        parser.add_argument(
            '--generate-template',
            action='store_true',
            help='Генерирует шаблон Excel файла с заголовками и примером данных',
        )

    def handle(self, *args, **options):
        generate_template = options.get('generate_template', False)
        
        if generate_template:
            self._generate_template(options['file_path'])
            return
        
        file_path = options['file_path']
        dry_run = options['dry_run']
        skip_errors = options['skip_errors']
        
        # Разрешаем путь относительно корня проекта (родительская папка backend)
        if not os.path.isabs(file_path):
            # Если путь относительный, разрешаем относительно корня проекта
            project_root = settings.BASE_DIR.parent  # Родитель backend (корень проекта)
            # Преобразуем Path в строку и нормализуем разделители
            file_path = str(project_root / file_path)
        
        # Нормализуем путь (для Windows и Linux)
        file_path = os.path.normpath(file_path)
        
        self.stdout.write(f'Импорт водителей из {file_path}...')
        if dry_run:
            self.stdout.write(self.style.WARNING('Режим валидации (dry-run): водители не будут созданы'))
        self.stdout.write('=' * 60)
        
        try:
            # Открываем Excel файл
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Определяем заголовки (первая строка)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    value = str(cell.value).strip().lower()
                else:
                    value = ''
                headers.append(value)
            
            self.stdout.write(f'Найдено колонок: {len(headers)}')
            self.stdout.write(f'Заголовки: {", ".join(headers)}')
            self.stdout.write('')
            
            # Маппинг возможных названий колонок (по полям из формы "Добавить водителя")
            column_mapping = {
                'name': ['name', 'имя', 'фио', 'full_name', 'driver_name'],
                'phone': ['phone', 'телефон', 'phone_number', 'номер телефона'],
                'email': ['email', 'почта', 'e-mail'],
                'password': ['password', 'пароль', 'pass'],
                'region_id': ['region_id', 'id региона', 'region id', 'регион id'],
                'region_title': ['region_title', 'название региона', 'регион', 'region'],
                'car_model': ['car_model', 'модель машины', 'автомобиль', 'машина', 'car', 'машина'],
                'plate_number': ['plate_number', 'номер машины', 'гос номер', 'plate', 'гос. номер', 'госномер'],
                'capacity': ['capacity', 'вместимость', 'мест', 'seats'],
                'is_online': ['is_online', 'онлайн', 'online', 'водитель онлайн', 'driver online'],
                'rating': ['rating', 'рейтинг']
            }
            
            # Создаем маппинг индексов колонок (с поддержкой частичных совпадений)
            column_indices = {}
            for key, possible_names in column_mapping.items():
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    # Проверяем точное совпадение
                    if header in possible_names:
                        column_indices[key] = idx
                        break
                    # Проверяем частичное совпадение (содержится в заголовке)
                    for name in possible_names:
                        if name in header or header in name:
                            column_indices[key] = idx
                            break
                    if key in column_indices:
                        break
            
            self.stdout.write('Маппинг колонок:')
            for key, idx in column_indices.items():
                self.stdout.write(f'  {key} -> колонка {idx + 1} ({headers[idx]})')
            self.stdout.write('')
            
            # Проверяем обязательные поля (соответствуют форме "Добавить водителя")
            required_fields = ['name', 'phone', 'password', 'car_model', 'plate_number', 'capacity']
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                self.stdout.write(self.style.ERROR(f'Отсутствуют обязательные колонки: {", ".join(missing_fields)}'))
                return
            
            # Проверяем регион (обязательное поле, может быть region_id или region_title)
            if 'region_id' not in column_indices and 'region_title' not in column_indices:
                self.stdout.write(self.style.ERROR('Отсутствует информация о регионе (нужна колонка region_id или region_title)'))
                return
            
            # Обрабатываем строки
            success_count = 0
            failed_count = 0
            updated_count = 0
            created_count = 0
            errors = []
            
            for row_num in range(2, sheet.max_row + 1):  # Начинаем с 2-й строки
                try:
                    row_data = {}
                    for key, idx in column_indices.items():
                        cell_value = sheet.cell(row=row_num, column=idx + 1).value
                        if cell_value is not None:
                            row_data[key] = str(cell_value).strip()
                        else:
                            row_data[key] = ''
                    
                    # Валидация и создание/обновление водителя
                    driver_data = self._parse_driver_row(row_data, row_num)
                    
                    if dry_run:
                        success_count += 1
                        continue
                    
                    # Создаем или обновляем пользователя и водителя
                    phone = driver_data['phone']
                    password = driver_data['password']
                    is_online = driver_data.get('is_online', False)
                    
                    # Проверяем, существует ли пользователь с таким телефоном
                    try:
                        user = User.objects.get(phone=phone)
                        # Обновляем существующего пользователя
                        if driver_data.get('email'):
                            user.email = driver_data['email']
                        if password:  # Обновляем пароль если указан
                            user.set_password(password)
                        user.role = 'driver'  # Убеждаемся, что роль установлена
                        user.save()
                        user_created = False
                    except User.DoesNotExist:
                        # Создаем нового пользователя
                        username = f'driver_{phone.replace("+", "").replace(" ", "").replace("-", "")}'
                        user = User.objects.create_user(
                            username=username,
                            phone=phone,
                            email=driver_data.get('email', ''),
                            password=password,
                            role='driver'
                        )
                        user_created = True
                    
                    # Создаем или обновляем водителя
                    driver, driver_created = Driver.objects.get_or_create(
                        user=user,
                        defaults={
                            'name': driver_data['name'],
                            'region': driver_data['region'],
                            'car_model': driver_data['car_model'],
                            'plate_number': driver_data['plate_number'],
                            'capacity': driver_data['capacity'],
                            'rating': driver_data.get('rating', 5.0),
                            'is_online': is_online,
                            'status': DriverStatus.ONLINE_IDLE if is_online else DriverStatus.OFFLINE,
                        }
                    )
                    
                    if not driver_created:
                        # Обновляем существующего водителя
                        updated = False
                        if driver.name != driver_data['name']:
                            driver.name = driver_data['name']
                            updated = True
                        if driver.region != driver_data['region']:
                            driver.region = driver_data['region']
                            updated = True
                        if driver.car_model != driver_data['car_model']:
                            driver.car_model = driver_data['car_model']
                            updated = True
                        if driver.plate_number != driver_data['plate_number']:
                            driver.plate_number = driver_data['plate_number']
                            updated = True
                        if driver.capacity != driver_data['capacity']:
                            driver.capacity = driver_data['capacity']
                            updated = True
                        if driver.is_online != is_online:
                            driver.is_online = is_online
                            # Обновляем статус в зависимости от is_online
                            driver.status = DriverStatus.ONLINE_IDLE if is_online else DriverStatus.OFFLINE
                            updated = True
                        if 'rating' in driver_data and driver.rating != driver_data['rating']:
                            driver.rating = driver_data['rating']
                            updated = True
                        
                        if updated:
                            driver.save()
                            updated_count += 1
                            self.stdout.write(self.style.WARNING(f'  Строка {row_num}: Водитель {driver.name} обновлен'))
                        else:
                            success_count += 1
                    else:
                        created_count += 1
                        success_count += 1
                        self.stdout.write(self.style.SUCCESS(f'  Строка {row_num}: Водитель {driver.name} создан'))
                        
                except Exception as e:
                    failed_count += 1
                    error_msg = f'Строка {row_num}: {str(e)}'
                    errors.append({'row': row_num, 'message': str(e)})
                    self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                    
                    if not skip_errors:
                        self.stdout.write(self.style.ERROR(f'Импорт остановлен из-за ошибки в строке {row_num}'))
                        break
            
            # Итоговый отчет
            self.stdout.write('')
            self.stdout.write('=' * 60)
            self.stdout.write(self.style.SUCCESS(f'Успешно обработано: {success_count}'))
            if not dry_run:
                self.stdout.write(f'  Создано: {created_count}')
                self.stdout.write(f'  Обновлено: {updated_count}')
            self.stdout.write(self.style.ERROR(f'Ошибок: {failed_count}'))
            
            if errors and (skip_errors or dry_run):
                self.stdout.write('')
                self.stdout.write('Детали ошибок:')
                for error in errors[:20]:
                    self.stdout.write(f"  Строка {error['row']}: {error['message']}")
                if len(errors) > 20:
                    self.stdout.write(f'  ... и еще {len(errors) - 20} ошибок')
                    
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Ошибка импорта: {str(e)}'))
            logger.error(f'Ошибка импорта водителей из Excel: {e}', exc_info=True)
    
    def _parse_driver_row(self, row_data: dict, row_num: int) -> dict:
        """Парсит строку Excel и возвращает данные для создания/обновления водителя"""
        # Обязательные поля (соответствуют форме "Добавить водителя")
        name = row_data.get('name', '').strip()
        phone = row_data.get('phone', '').strip()
        password = row_data.get('password', '').strip()
        
        if not name:
            raise ValueError('Не указано имя водителя')
        if not phone:
            raise ValueError('Не указан телефон водителя')
        if not password:
            raise ValueError('Не указан пароль водителя')
        
        # Валидация пароля (минимум 8 символов)
        if len(password) < 8:
            raise ValueError('Пароль должен содержать минимум 8 символов')
        
        # Нормализация телефона (удаление пробелов, дефисов, скобок)
        phone_cleaned = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if not phone_cleaned.isdigit() or len(phone_cleaned) < 10:
            raise ValueError('Неверный формат телефона')
        
        # Регион (обязательное поле)
        region_id = row_data.get('region_id', '').strip()
        region_title = row_data.get('region_title', '').strip()
        
        region = None
        if region_id:
            try:
                region = Region.objects.get(id=region_id)
            except Region.DoesNotExist:
                raise ValueError(f'Регион с ID {region_id} не найден')
        elif region_title:
            # Ищем по названию (может быть неточное совпадение)
            region = Region.objects.filter(title__icontains=region_title).first()
            if not region:
                raise ValueError(f'Регион "{region_title}" не найден')
        else:
            raise ValueError('Не указан регион (region_id или region_title)')
        
        # Автомобиль (обязательные поля)
        car_model = row_data.get('car_model', '').strip()
        plate_number = row_data.get('plate_number', '').strip()
        
        if not car_model:
            raise ValueError('Не указана модель машины')
        if not plate_number:
            raise ValueError('Не указан номер машины')
        
        # Вместимость (обязательное поле, по умолчанию 4)
        capacity_str = row_data.get('capacity', '4').strip()
        try:
            capacity = int(capacity_str) if capacity_str else 4
            if capacity < 1 or capacity > 20:
                raise ValueError('Вместимость должна быть от 1 до 20 мест')
        except ValueError as e:
            if 'Вместимость должна быть' in str(e):
                raise
            raise ValueError('Неверный формат вместимости (должно быть число)')
        
        # Водитель онлайн (опционально, по умолчанию False)
        is_online_str = row_data.get('is_online', '').strip().lower()
        is_online = False
        if is_online_str:
            is_online = is_online_str in ('true', '1', 'yes', 'да', 'д', 'онлайн', 'online')
        
        # Рейтинг (опционально, по умолчанию 5.0)
        rating_str = row_data.get('rating', '').strip()
        rating = 5.0
        if rating_str:
            try:
                rating = float(rating_str)
                if rating < 0 or rating > 5:
                    raise ValueError('Рейтинг должен быть от 0 до 5')
            except ValueError as e:
                if 'Рейтинг должен быть' in str(e):
                    raise
                pass  # Если не удалось распарсить, используем значение по умолчанию
        
        # Email (опционально)
        email = row_data.get('email', '').strip()
        if email and '@' not in email:
            raise ValueError('Неверный формат email')
        
        return {
            'name': name,
            'phone': phone,
            'password': password,
            'email': email,
            'region': region,
            'car_model': car_model,
            'plate_number': plate_number,
            'capacity': capacity,
            'is_online': is_online,
            'rating': rating
        }
    
    def _generate_template(self, file_path=None):
        """Генерирует шаблон Excel файла для импорта водителей"""
        # Определяем путь к шаблону
        if not file_path or file_path == 'data/водитель (1).xlsx':  # Значение по умолчанию
            project_root = settings.BASE_DIR.parent
            file_path = str(project_root / 'data' / 'шаблон_импорт_водителей.xlsx')
        else:
            # Разрешаем путь относительно корня проекта
            if not os.path.isabs(file_path):
                project_root = settings.BASE_DIR.parent
                file_path = str(project_root / file_path)
            file_path = os.path.normpath(file_path)
        
        # Создаем директорию, если её нет
        directory = os.path.dirname(file_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        
        # Создаем новую рабочую книгу
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Водители'
        
        # Определяем заголовки колонок (русские названия для удобства)
        headers = [
            'Имя',
            'Телефон',
            'Email',
            'Пароль',
            'Регион',
            'Машина',
            'Гос. номер',
            'Вместимость',
            'Водитель онлайн',
            'Рейтинг'
        ]
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            # Делаем заголовки жирными
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Добавляем пример данных
        # Получаем первый регион для примера
        first_region = Region.objects.first()
        region_title = first_region.title if first_region else 'Алматы'
        region_id = first_region.id if first_region else '1'
        
        example_row = [
            'Иванов Иван Иванович',
            '+7 777 123 4567',
            'driver1@invotaxi.kz',
            'password123',
            region_title,  # Используем название региона
            'Toyota Camry',
            'A 123 BC 02',
            '4',
            'Нет',
            '5.0'
        ]
        
        for col_idx, value in enumerate(example_row, start=1):
            sheet.cell(row=2, column=col_idx).value = value
        
        # Добавляем второй пример (опциональные поля)
        example_row2 = [
            'Петров Петр Петрович',
            '+7 777 765 4321',
            'driver2@invotaxi.kz',
            'securepass123',
            region_title,
            'Hyundai Sonata',
            'B 456 DE 05',
            '4',
            'Да',
            '4.5'
        ]
        
        for col_idx, value in enumerate(example_row2, start=1):
            sheet.cell(row=3, column=col_idx).value = value
        
        # Настраиваем ширину колонок
        column_widths = {
            'A': 25,  # Имя
            'B': 18,  # Телефон
            'C': 25,  # Email
            'D': 20,  # Пароль
            'E': 20,  # Регион
            'F': 18,  # Машина
            'G': 15,  # Гос. номер
            'H': 12,  # Вместимость
            'I': 18,  # Водитель онлайн
            'J': 10   # Рейтинг
        }
        
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        # Добавляем подсказки на втором листе
        sheet2 = workbook.create_sheet(title='Инструкция')
        instructions = [
            ['ПОЛЕ', 'ОБЯЗАТЕЛЬНО', 'ОПИСАНИЕ', 'ПРИМЕР'],
            ['Имя', 'Да', 'Полное имя водителя', 'Иванов Иван Иванович'],
            ['Телефон', 'Да', 'Телефон в формате +7...', '+7 777 123 4567'],
            ['Email', 'Нет', 'Email адрес (опционально)', 'driver@invotaxi.kz'],
            ['Пароль', 'Да', 'Минимум 8 символов', 'password123'],
            ['Регион', 'Да', 'Название региона или ID региона', 'Алматы'],
            ['Машина', 'Да', 'Модель автомобиля', 'Toyota Camry'],
            ['Гос. номер', 'Да', 'Государственный номер', 'A 123 BC 02'],
            ['Вместимость', 'Да', 'Количество мест (1-20)', '4'],
            ['Водитель онлайн', 'Нет', 'Да/Нет или True/False', 'Нет'],
            ['Рейтинг', 'Нет', 'Рейтинг от 0 до 5', '5.0'],
            ['', '', '', ''],
            ['ПРИМЕЧАНИЯ:', '', '', ''],
            ['1. Все обязательные поля должны быть заполнены', '', '', ''],
            ['2. Телефон должен быть уникальным', '', '', ''],
            ['3. Пароль должен содержать минимум 8 символов', '', '', ''],
            ['4. Для региона можно использовать название или ID', '', '', ''],
            ['5. Вместимость должна быть от 1 до 20 мест', '', '', ''],
            ['6. Для "Водитель онлайн" используйте: Да/Нет, True/False, 1/0', '', '', ''],
        ]
        
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet2.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Форматируем заголовок инструкции
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx == 12:  # Заголовок "ПРИМЕЧАНИЯ"
                    if col_idx == 1:  # Только для первой колонки
                        cell.font = Font(bold=True, size=11)
        
        # Настраиваем ширину колонок для инструкции
        sheet2.column_dimensions['A'].width = 20
        sheet2.column_dimensions['B'].width = 12
        sheet2.column_dimensions['C'].width = 50
        sheet2.column_dimensions['D'].width = 25
        
        # Сохраняем файл
        workbook.save(file_path)
        
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'[OK] Шаблон создан: {file_path}'))
        self.stdout.write('')
        self.stdout.write('Шаблон содержит:')
        self.stdout.write('  - Лист "Водители": заголовки колонок и 2 примера данных')
        self.stdout.write('  - Лист "Инструкция": описание полей и требования')
        self.stdout.write('')
        
        if first_region:
            self.stdout.write(self.style.WARNING(
                f'В примерах использован регион: {region_title} (ID: {region_id})'
            ))
            self.stdout.write('Доступные регионы:')
            for region in Region.objects.all()[:10]:
                self.stdout.write(f'  - {region.title} (ID: {region.id})')
            if Region.objects.count() > 10:
                self.stdout.write(f'  ... и еще {Region.objects.count() - 10} регионов')
        else:
            self.stdout.write(self.style.ERROR(
                'ВНИМАНИЕ: В системе нет регионов! Создайте регионы перед импортом водителей.'
            ))
