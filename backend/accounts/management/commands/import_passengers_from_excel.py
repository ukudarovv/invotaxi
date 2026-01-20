"""
Management command для импорта пассажиров из Excel файла
"""
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.conf import settings
from accounts.models import Passenger
from regions.models import Region
import openpyxl
import logging
import os

logger = logging.getLogger(__name__)
User = get_user_model()


class Command(BaseCommand):
    help = '''
    Импортирует пассажиров из Excel файла.
    
    Поддерживаемые поля:
    Обязательные:
    - Имя* (full_name, имя, фио, name)
    - Телефон* (phone, телефон)
    - Регион* (region_id, region_title, регион, id региона, название региона)
    - Категория инвалидности* (disability_category, категория) - I группа, II группа, III группа, Ребенок-инвалид
    
    Опциональные:
    - Email (email, почта)
    - Разрешено сопровождение (allowed_companion, сопровождение, companion) - Да/Нет, True/False, 1/0
    
    Пример использования:
    python manage.py import_passengers_from_excel "data/passengers.xlsx" --skip-errors
    
    Для генерации шаблона:
    python manage.py import_passengers_from_excel --generate-template
    python manage.py import_passengers_from_excel --generate-template "data/шаблон_пассажиры.xlsx"
    '''

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            nargs='?',
            default='data/passengers.xlsx',
            help='Путь к Excel файлу (по умолчанию: data/passengers.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Только валидация, без создания пассажиров',
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Пропускать строки с ошибками и продолжать',
        )
        parser.add_argument(
            '--generate-template',
            action='store_true',
            help='Генерирует шаблон Excel файла с заголовками и примером данных',
        )

    def handle(self, *args, **options):
        generate_template = options.get('generate_template', False)
        
        if generate_template:
            self._generate_template(options.get('file_path'))
            return
        
        file_path = options['file_path']
        dry_run = options['dry_run']
        skip_errors = options['skip_errors']
        
        # Разрешаем путь относительно корня проекта
        if not os.path.isabs(file_path):
            project_root = settings.BASE_DIR.parent
            file_path = str(project_root / file_path)
        
        file_path = os.path.normpath(file_path)
        
        self.stdout.write(f'Импорт пассажиров из {file_path}...')
        if dry_run:
            self.stdout.write(self.style.WARNING('Режим валидации (dry-run): пассажиры не будут созданы'))
        self.stdout.write('=' * 60)
        
        try:
            # Открываем Excel файл
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Определяем заголовки (первая строка)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    value = str(cell.value).strip().lower()
                else:
                    value = ''
                headers.append(value)
            
            self.stdout.write(f'Найдено колонок: {len(headers)}')
            self.stdout.write(f'Заголовки: {", ".join(headers)}')
            self.stdout.write('')
            
            # Маппинг возможных названий колонок
            column_mapping = {
                'full_name': ['full_name', 'имя', 'фио', 'name', 'passenger_name', 'пассажир'],
                'phone': ['phone', 'телефон', 'phone_number', 'номер телефона', 'тел'],
                'email': ['email', 'почта', 'e-mail'],
                'region_id': ['region_id', 'id региона', 'region id', 'регион id'],
                'region_title': ['region_title', 'название региона', 'регион', 'region'],
                'disability_category': ['disability_category', 'категория', 'категория инвалидности', 'disability'],
                'allowed_companion': ['allowed_companion', 'сопровождение', 'companion', 'разрешено сопровождение', 'сопр'],
            }
            
            # Создаем маппинг индексов колонок
            column_indices = {}
            for key, possible_names in column_mapping.items():
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    # Проверяем точное совпадение
                    if header in [name.lower() for name in possible_names]:
                        column_indices[key] = idx
                        break
                    # Проверяем частичное совпадение
                    for name in possible_names:
                        if name.lower() in header or header in name.lower():
                            column_indices[key] = idx
                            break
                    if key in column_indices:
                        break
            
            self.stdout.write('Маппинг колонок:')
            for key, idx in column_indices.items():
                self.stdout.write(f'  {key} -> колонка {idx + 1} ({headers[idx]})')
            self.stdout.write('')
            
            # Проверяем обязательные поля
            required_fields = ['full_name', 'phone', 'disability_category']
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                self.stdout.write(self.style.ERROR(f'Отсутствуют обязательные колонки: {", ".join(missing_fields)}'))
                return
            
            # Проверяем регион (обязательное поле, может быть region_id или region_title)
            if 'region_id' not in column_indices and 'region_title' not in column_indices:
                self.stdout.write(self.style.ERROR('Отсутствует информация о регионе (нужна колонка region_id или region_title)'))
                return
            
            # Обрабатываем строки
            success_count = 0
            failed_count = 0
            updated_count = 0
            created_count = 0
            errors = []
            
            for row_num in range(2, sheet.max_row + 1):  # Начинаем с 2-й строки
                try:
                    row_data = {}
                    for key, idx in column_indices.items():
                        cell_value = sheet.cell(row=row_num, column=idx + 1).value
                        if cell_value is not None:
                            row_data[key] = str(cell_value).strip()
                        else:
                            row_data[key] = ''
                    
                    # Валидация и создание/обновление пассажира
                    passenger_data = self._parse_passenger_row(row_data, row_num)
                    
                    if dry_run:
                        success_count += 1
                        continue
                    
                    # Создаем или обновляем пользователя и пассажира
                    phone = passenger_data['phone']
                    email = passenger_data.get('email', '')
                    
                    # Проверяем, существует ли пользователь с таким телефоном
                    try:
                        user = User.objects.get(phone=phone)
                        # Обновляем существующего пользователя
                        if email:
                            user.email = email
                        user.role = 'passenger'  # Убеждаемся, что роль установлена
                        user.save()
                        user_created = False
                    except User.DoesNotExist:
                        # Создаем нового пользователя
                        username = f'passenger_{phone.replace("+", "").replace(" ", "").replace("-", "").replace("(", "").replace(")", "")}'
                        user = User.objects.create_user(
                            username=username,
                            phone=phone,
                            email=email,
                            password=None,  # Пароль не требуется для пассажиров
                            role='passenger'
                        )
                        user_created = True
                    
                    # Создаем или обновляем пассажира
                    passenger, passenger_created = Passenger.objects.get_or_create(
                        user=user,
                        defaults={
                            'full_name': passenger_data['full_name'],
                            'region': passenger_data['region'],
                            'disability_category': passenger_data['disability_category'],
                            'allowed_companion': passenger_data.get('allowed_companion', False),
                        }
                    )
                    
                    if not passenger_created:
                        # Обновляем существующего пассажира
                        updated = False
                        if passenger.full_name != passenger_data['full_name']:
                            passenger.full_name = passenger_data['full_name']
                            updated = True
                        if passenger.region != passenger_data['region']:
                            passenger.region = passenger_data['region']
                            updated = True
                        if passenger.disability_category != passenger_data['disability_category']:
                            passenger.disability_category = passenger_data['disability_category']
                            updated = True
                        if passenger.allowed_companion != passenger_data.get('allowed_companion', False):
                            passenger.allowed_companion = passenger_data.get('allowed_companion', False)
                            updated = True
                        
                        if updated:
                            passenger.save()
                            updated_count += 1
                            self.stdout.write(self.style.WARNING(f'  Строка {row_num}: Пассажир {passenger.full_name} обновлен'))
                        else:
                            success_count += 1
                    else:
                        created_count += 1
                        success_count += 1
                        self.stdout.write(self.style.SUCCESS(f'  Строка {row_num}: Пассажир {passenger.full_name} создан'))
                        
                except Exception as e:
                    failed_count += 1
                    error_msg = f'Строка {row_num}: {str(e)}'
                    errors.append({'row': row_num, 'message': str(e)})
                    self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                    
                    if not skip_errors:
                        self.stdout.write(self.style.ERROR(f'Импорт остановлен из-за ошибки в строке {row_num}'))
                        break
            
            # Итоговый отчет
            self.stdout.write('')
            self.stdout.write('=' * 60)
            self.stdout.write(self.style.SUCCESS(f'Успешно обработано: {success_count}'))
            if not dry_run:
                self.stdout.write(f'  Создано: {created_count}')
                self.stdout.write(f'  Обновлено: {updated_count}')
            self.stdout.write(self.style.ERROR(f'Ошибок: {failed_count}'))
            
            if errors and (skip_errors or dry_run):
                self.stdout.write('')
                self.stdout.write('Детали ошибок:')
                for error in errors[:20]:
                    self.stdout.write(f"  Строка {error['row']}: {error['message']}")
                if len(errors) > 20:
                    self.stdout.write(f'  ... и еще {len(errors) - 20} ошибок')
                    
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Ошибка импорта: {str(e)}'))
            logger.error(f'Ошибка импорта пассажиров из Excel: {e}', exc_info=True)
    
    def _parse_passenger_row(self, row_data: dict, row_num: int) -> dict:
        """Парсит строку Excel и возвращает данные для создания/обновления пассажира"""
        # Обязательные поля
        full_name = row_data.get('full_name', '').strip()
        phone = row_data.get('phone', '').strip()
        disability_category = row_data.get('disability_category', '').strip()
        
        if not full_name:
            raise ValueError('Не указано имя пассажира')
        if not phone:
            raise ValueError('Не указан телефон пассажира')
        if not disability_category:
            raise ValueError('Не указана категория инвалидности')
        
        # Валидация телефона
        phone_cleaned = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if not phone_cleaned.isdigit() or len(phone_cleaned) < 10:
            raise ValueError('Неверный формат телефона')
        
        # Валидация категории инвалидности
        valid_categories = ['I группа', 'II группа', 'III группа', 'Ребенок-инвалид']
        if disability_category not in valid_categories:
            raise ValueError(f'Неверная категория инвалидности. Допустимые значения: {", ".join(valid_categories)}')
        
        # Регион (обязательное поле)
        region_id = row_data.get('region_id', '').strip()
        region_title = row_data.get('region_title', '').strip()
        
        region = None
        if region_id:
            try:
                region = Region.objects.get(id=region_id)
            except Region.DoesNotExist:
                raise ValueError(f'Регион с ID {region_id} не найден')
        elif region_title:
            # Ищем по названию (может быть неточное совпадение)
            region = Region.objects.filter(title__icontains=region_title).first()
            if not region:
                raise ValueError(f'Регион "{region_title}" не найден')
        else:
            raise ValueError('Не указан регион (region_id или region_title)')
        
        # Email (опционально)
        email = row_data.get('email', '').strip()
        if email and '@' not in email:
            raise ValueError('Неверный формат email')
        
        # Разрешено сопровождение (опционально, по умолчанию False)
        allowed_companion_str = row_data.get('allowed_companion', '').strip().lower()
        allowed_companion = False
        if allowed_companion_str:
            allowed_companion = allowed_companion_str in ('true', '1', 'yes', 'да', 'д')
        
        return {
            'full_name': full_name,
            'phone': phone,
            'email': email,
            'region': region,
            'disability_category': disability_category,
            'allowed_companion': allowed_companion
        }
    
    def _generate_template(self, file_path=None):
        """Генерирует шаблон Excel файла для импорта пассажиров"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Определяем путь к шаблону
        if not file_path or file_path == 'data/passengers.xlsx':  # Значение по умолчанию
            project_root = settings.BASE_DIR.parent
            file_path = str(project_root / 'data' / 'шаблон_импорт_пассажиров.xlsx')
        else:
            # Разрешаем путь относительно корня проекта
            if not os.path.isabs(file_path):
                project_root = settings.BASE_DIR.parent
                file_path = str(project_root / file_path)
            file_path = os.path.normpath(file_path)
        
        # Создаем директорию, если её нет
        directory = os.path.dirname(file_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        
        # Создаем новую рабочую книгу
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Пассажиры'
        
        # Определяем заголовки колонок
        headers = [
            'Имя',
            'Телефон',
            'Email',
            'Регион',
            'Категория инвалидности',
            'Разрешено сопровождение'
        ]
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Добавляем пример данных
        first_region = Region.objects.first()
        region_title = first_region.title if first_region else 'Алматы'
        
        example_row = [
            'Иванов Иван Иванович',
            '+7 777 123 4567',
            'passenger1@invotaxi.kz',
            region_title,
            'III группа',
            'Нет'
        ]
        
        for col_idx, value in enumerate(example_row, start=1):
            sheet.cell(row=2, column=col_idx).value = value
        
        # Второй пример
        example_row2 = [
            'Петрова Мария Сергеевна',
            '+7 777 765 4321',
            'passenger2@invotaxi.kz',
            region_title,
            'I группа',
            'Да'
        ]
        
        for col_idx, value in enumerate(example_row2, start=1):
            sheet.cell(row=3, column=col_idx).value = value
        
        # Настраиваем ширину колонок
        column_widths = {
            'A': 25, 'B': 18, 'C': 25, 'D': 20, 'E': 25, 'F': 22
        }
        for col_letter, width in column_widths.items():
            sheet.column_dimensions[col_letter].width = width
        
        # Добавляем инструкцию на второй лист
        sheet2 = workbook.create_sheet(title='Инструкция')
        instructions = [
            ['ПОЛЕ', 'ОБЯЗАТЕЛЬНО', 'ОПИСАНИЕ', 'ПРИМЕР'],
            ['Имя', 'Да', 'Полное имя пассажира', 'Иванов Иван Иванович'],
            ['Телефон', 'Да', 'Телефон в формате +7...', '+7 777 123 4567'],
            ['Email', 'Нет', 'Email адрес (опционально)', 'passenger@invotaxi.kz'],
            ['Регион', 'Да', 'Название региона или ID региона', 'Алматы'],
            ['Категория инвалидности', 'Да', 'I группа, II группа, III группа, Ребенок-инвалид', 'III группа'],
            ['Разрешено сопровождение', 'Нет', 'Да/Нет или True/False', 'Нет'],
            ['', '', '', ''],
            ['ПРИМЕЧАНИЯ:', '', '', ''],
            ['1. Все обязательные поля должны быть заполнены', '', '', ''],
            ['2. Телефон должен быть уникальным', '', '', ''],
            ['3. Для региона можно использовать название или ID', '', '', ''],
            ['4. Категория инвалидности должна быть одной из: I группа, II группа, III группа, Ребенок-инвалид', '', '', ''],
            ['5. Для "Разрешено сопровождение" используйте: Да/Нет, True/False, 1/0', '', '', ''],
        ]
        
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet2.cell(row=row_idx, column=col_idx)
                cell.value = value
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif row_idx == 9 and col_idx == 1:
                    cell.font = Font(bold=True, size=11)
        
        sheet2.column_dimensions['A'].width = 25
        sheet2.column_dimensions['B'].width = 12
        sheet2.column_dimensions['C'].width = 60
        sheet2.column_dimensions['D'].width = 30
        
        # Сохраняем файл
        workbook.save(file_path)
        
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(f'[OK] Шаблон создан: {file_path}'))
        self.stdout.write('')
        self.stdout.write('Шаблон содержит:')
        self.stdout.write('  - Лист "Пассажиры": заголовки колонок и 2 примера данных')
        self.stdout.write('  - Лист "Инструкция": описание полей и требования')
        self.stdout.write('')
        
        if first_region:
            self.stdout.write(self.style.WARNING(
                f'В примерах использован регион: {region_title} (ID: {first_region.id})'
            ))
            self.stdout.write('Доступные регионы:')
            for region in Region.objects.all()[:10]:
                self.stdout.write(f'  - {region.title} (ID: {region.id})')
            if Region.objects.count() > 10:
                self.stdout.write(f'  ... и еще {Region.objects.count() - 10} регионов')
        else:
            self.stdout.write(self.style.ERROR(
                'ВНИМАНИЕ: В системе нет регионов! Создайте регионы перед импортом пассажиров.'
            ))
