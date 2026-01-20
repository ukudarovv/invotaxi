from rest_framework import status, viewsets, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.db import models
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Passenger, Driver, UserActivityLog
from regions.models import Region
from .serializers import (
    UserSerializer, PassengerSerializer, DriverSerializer, DriverCreateSerializer,
    PhoneLoginSerializer, VerifyOTPSerializer, EmailPasswordLoginSerializer,
    AdminUserSerializer, AdminUserCreateSerializer, AdminUserUpdateSerializer,
    PasswordResetSerializer, BulkActionSerializer, UserActivityLogSerializer
)
from django.contrib.auth import authenticate
from .services import OTPService, UserService

User = get_user_model()


class AuthViewSet(viewsets.ViewSet):
    """ViewSet для аутентификации"""
    permission_classes = [AllowAny]

    @action(detail=False, methods=['post'], url_path='phone-login')
    def phone_login(self, request):
        """Запрос OTP кода"""
        serializer = PhoneLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        phone = serializer.validated_data['phone']
        otp = OTPService.create_otp(phone)

        return Response({
            'message': 'OTP код отправлен',
            'expires_at': otp.expires_at
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='verify-otp')
    def verify_otp(self, request):
        """Проверка OTP и выдача токена"""
        serializer = VerifyOTPSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        phone = serializer.validated_data['phone']
        code = serializer.validated_data['code']

        is_valid, otp = OTPService.verify_otp(phone, code)

        if not is_valid:
            return Response(
                {'error': 'Неверный или истекший код'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Получаем или создаем пользователя
        user, created = OTPService.get_or_create_user(phone)

        # Генерируем JWT токены
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)

        # Определяем роль и дополнительные данные
        role = UserService.get_user_role(user)
        response_data = {
            'access': access_token,
            'refresh': str(refresh),
            'user': UserSerializer(user).data,
            'role': role,
        }

        # Добавляем данные пассажира или водителя
        if role == 'passenger' and hasattr(user, 'passenger'):
            response_data['passenger'] = PassengerSerializer(user.passenger).data
            response_data['passenger_id'] = user.passenger.id
        elif role == 'driver' and hasattr(user, 'driver'):
            response_data['driver'] = DriverSerializer(user.driver).data
            response_data['driver_id'] = user.driver.id

        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='email-login')
    def email_login(self, request):
        """Вход по email и паролю для админ-панели"""
        serializer = EmailPasswordLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email']
        password = serializer.validated_data['password']

        # Пытаемся найти пользователя по email
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response(
                {'error': 'Неверный email или пароль'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Проверяем пароль
        if not user.check_password(password):
            return Response(
                {'error': 'Неверный email или пароль'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        # Проверяем, что пользователь является staff (админ, диспетчер, оператор)
        if not user.is_staff:
            return Response(
                {'error': 'Доступ разрешен только для администраторов'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Генерируем JWT токены
        refresh = RefreshToken.for_user(user)
        access_token = str(refresh.access_token)

        # Определяем роль пользователя
        # Можно использовать группы Django или флаги
        role = 'admin'  # По умолчанию
        if user.groups.filter(name='dispatcher').exists():
            role = 'dispatcher'
        elif user.groups.filter(name='operator').exists():
            role = 'operator'

        response_data = {
            'access': access_token,
            'refresh': str(refresh),
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'phone': user.phone,
                'role': role,
            },
            'role': role,
        }

        return Response(response_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='refresh')
    def refresh_token(self, request):
        """Обновление access токена"""
        refresh_token = request.data.get('refresh')
        if not refresh_token:
            return Response(
                {'error': 'Требуется refresh токен'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            refresh = RefreshToken(refresh_token)
            access_token = str(refresh.access_token)
            return Response({
                'access': access_token,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'error': 'Неверный или истекший refresh токен'},
                status=status.HTTP_401_UNAUTHORIZED
            )

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def logout(self, request):
        """Выход из системы"""
        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
        except Exception:
            pass

        return Response({'message': 'Выход выполнен'}, status=status.HTTP_200_OK)


class PassengerViewSet(viewsets.ModelViewSet):
    """ViewSet для пассажиров"""
    queryset = Passenger.objects.select_related('user', 'region').all()
    serializer_class = PassengerSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        # Пассажиры могут видеть только свой профиль
        if hasattr(user, 'passenger'):
            return Passenger.objects.filter(id=user.passenger.id)
        # Админы и водители могут видеть всех
        return super().get_queryset()

    def create(self, request, *args, **kwargs):
        """Создание пассажира - только для админов"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут создавать пассажиров')
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        """Обновление пассажира - только для админов"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут обновлять пассажиров')
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Удаление пассажира - только для админов"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут удалять пассажиров')
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """Скачивание шаблона Excel для импорта пассажиров"""
        try:
            # Проверяем права доступа (только админы могут скачивать шаблон)
            if not request.user.is_staff:
                raise PermissionDenied('Только администраторы могут скачивать шаблон')
            
            # Генерируем шаблон в памяти
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Пассажиры'
            
            # Определяем заголовки колонок
            headers = [
                'Имя',
                'Телефон',
                'Email',
                'Регион',
                'Категория инвалидности',
                'Разрешено сопровождение'
            ]
            
            # Записываем заголовки
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем пример данных
            first_region = Region.objects.first()
            region_title = first_region.title if first_region else 'Алматы'
            
            example_row = [
                'Иванов Иван Иванович',
                '+7 777 123 4567',
                'passenger1@invotaxi.kz',
                region_title,
                'III группа',
                'Нет'
            ]
            
            for col_idx, value in enumerate(example_row, start=1):
                sheet.cell(row=2, column=col_idx).value = value
            
            # Второй пример
            example_row2 = [
                'Петрова Мария Сергеевна',
                '+7 777 765 4321',
                'passenger2@invotaxi.kz',
                region_title,
                'I группа',
                'Да'
            ]
            
            for col_idx, value in enumerate(example_row2, start=1):
                sheet.cell(row=3, column=col_idx).value = value
            
            # Настраиваем ширину колонок
            column_widths = {
                'A': 25, 'B': 18, 'C': 25, 'D': 20, 'E': 25, 'F': 22
            }
            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width
            
            # Добавляем инструкцию на второй лист
            sheet2 = workbook.create_sheet(title='Инструкция')
            instructions = [
                ['ПОЛЕ', 'ОБЯЗАТЕЛЬНО', 'ОПИСАНИЕ', 'ПРИМЕР'],
                ['Имя', 'Да', 'Полное имя пассажира', 'Иванов Иван Иванович'],
                ['Телефон', 'Да', 'Телефон в формате +7...', '+7 777 123 4567'],
                ['Email', 'Нет', 'Email адрес (опционально)', 'passenger@invotaxi.kz'],
                ['Регион', 'Да', 'Название региона или ID региона', 'Алматы'],
                ['Категория инвалидности', 'Да', 'I группа, II группа, III группа, Ребенок-инвалид', 'III группа'],
                ['Разрешено сопровождение', 'Нет', 'Да/Нет или True/False', 'Нет'],
                ['', '', '', ''],
                ['ПРИМЕЧАНИЯ:', '', '', ''],
                ['1. Все обязательные поля должны быть заполнены', '', '', ''],
                ['2. Телефон должен быть уникальным', '', '', ''],
                ['3. Для региона можно использовать название или ID', '', '', ''],
                ['4. Категория инвалидности должна быть одной из: I группа, II группа, III группа, Ребенок-инвалид', '', '', ''],
                ['5. Для "Разрешено сопровождение" используйте: Да/Нет, True/False, 1/0', '', '', ''],
            ]
            
            for row_idx, row_data in enumerate(instructions, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet2.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=12, color='FFFFFF')
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif row_idx == 9 and col_idx == 1:
                        cell.font = Font(bold=True, size=11)
            
            sheet2.column_dimensions['A'].width = 25
            sheet2.column_dimensions['B'].width = 12
            sheet2.column_dimensions['C'].width = 60
            sheet2.column_dimensions['D'].width = 30
            
            # Сохраняем в BytesIO
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            
            # Возвращаем файл как HTTP response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="шаблон_импорт_пассажиров.xlsx"'
            return response
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Ошибка генерации шаблона: {e}', exc_info=True)
            return Response(
                {'error': f'Ошибка генерации шаблона: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='import')
    def import_passengers(self, request):
        """Импорт пассажиров из Excel файла"""
        # Проверяем права доступа (только админы могут импортировать)
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут импортировать пассажиров')
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Файл не предоставлен'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Файл должен быть Excel файлом (.xlsx или .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Получаем опции из запроса
        skip_errors = request.data.get('skip_errors', 'false').lower() == 'true'
        dry_run = request.data.get('dry_run', 'false').lower() == 'true'
        
        # Используем команду импорта для обработки файла
        from accounts.management.commands.import_passengers_from_excel import Command as ImportCommand
        import tempfile
        import os
        
        # Сохраняем файл во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in csv_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Создаем экземпляр команды
            command = ImportCommand()
            
            # Перенаправляем вывод команды
            from io import StringIO
            import sys
            
            output_buffer = StringIO()
            original_stdout = sys.stdout
            sys.stdout = output_buffer
            
            try:
                # Вызываем обработку файла
                command.handle(
                    file_path=tmp_file_path,
                    dry_run=dry_run,
                    skip_errors=skip_errors
                )
            finally:
                sys.stdout = original_stdout
            
            output = output_buffer.getvalue()
            
            # Парсим вывод для получения статистики
            success_count = 0
            created_count = 0
            updated_count = 0
            failed_count = 0
            errors = []
            
            # Пытаемся извлечь статистику из вывода
            import re
            success_match = re.search(r'Успешно обработано: (\d+)', output)
            if success_match:
                success_count = int(success_match.group(1))
            
            created_match = re.search(r'Создано: (\d+)', output)
            if created_match:
                created_count = int(created_match.group(1))
            
            updated_match = re.search(r'Обновлено: (\d+)', output)
            if updated_match:
                updated_count = int(updated_match.group(1))
            
            failed_match = re.search(r'Ошибок: (\d+)', output)
            if failed_match:
                failed_count = int(failed_match.group(1))
            
            return Response({
                'success': True,
                'message': 'Импорт завершен',
                'statistics': {
                    'success_count': success_count,
                    'created_count': created_count,
                    'updated_count': updated_count,
                    'failed_count': failed_count,
                },
                'output': output if failed_count > 0 else None,  # Показываем вывод только при ошибках
            })
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            return Response(
                {
                    'success': False,
                    'error': str(e),
                    'traceback': error_trace if request.user.is_superuser else None
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Удаляем временный файл
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)


class DriverViewSet(viewsets.ModelViewSet):
    """ViewSet для водителей"""
    queryset = Driver.objects.select_related('user', 'region').all()
    serializer_class = DriverSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        """Выбор сериализатора в зависимости от действия"""
        if self.action == 'create':
            return DriverCreateSerializer
        return DriverSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()
        
        # Водители могут видеть только свой профиль
        if hasattr(user, 'driver'):
            return queryset.filter(id=user.driver.id)
        
        # Админы могут видеть всех с фильтрацией
        # Фильтрация по is_online
        is_online = self.request.query_params.get('is_online')
        if is_online is not None:
            queryset = queryset.filter(is_online=is_online.lower() == 'true')
        
        # Фильтрация по region_id
        region_id = self.request.query_params.get('region_id')
        if region_id:
            queryset = queryset.filter(region_id=region_id)
        
        # Поиск по name, car_model, plate_number, user__phone
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(name__icontains=search) |
                models.Q(car_model__icontains=search) |
                models.Q(plate_number__icontains=search) |
                models.Q(user__phone__icontains=search)
            )
        
        return queryset

    @action(detail=True, methods=['patch'], url_path='location')
    def update_location(self, request, pk=None):
        """Обновление позиции водителя"""
        driver = self.get_object()
        
        # Проверяем права доступа
        # Администраторы могут обновлять позицию любого водителя
        # Водители могут обновлять только свою позицию
        if hasattr(request.user, 'driver') and driver.id != request.user.driver.id:
            if not request.user.is_staff:
                return Response(
                    {'error': 'Нет доступа. Вы можете обновить только свою позицию'},
                    status=status.HTTP_403_FORBIDDEN
                )

        lat = request.data.get('lat')
        lon = request.data.get('lon')

        if lat is None or lon is None:
            return Response(
                {'error': 'Требуются lat и lon'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Валидация координат
        try:
            lat = float(lat)
            lon = float(lon)
            
            # Проверка диапазона координат (широта: -90 до 90, долгота: -180 до 180)
            if not (-90 <= lat <= 90):
                return Response(
                    {'error': 'Широта должна быть в диапазоне от -90 до 90'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            if not (-180 <= lon <= 180):
                return Response(
                    {'error': 'Долгота должна быть в диапазоне от -180 до 180'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Неверный формат координат. Требуются числа'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.utils import timezone
        driver.current_lat = lat
        driver.current_lon = lon
        driver.last_location_update = timezone.now()
        driver.save(update_fields=['current_lat', 'current_lon', 'last_location_update'])

        return Response(DriverSerializer(driver).data)

    @action(detail=True, methods=['patch'], url_path='online-status')
    def update_online_status(self, request, pk=None):
        """Обновление онлайн статуса водителя"""
        driver = self.get_object()
        
        # Проверяем права доступа
        if hasattr(request.user, 'driver') and driver.id != request.user.driver.id:
            return Response(
                {'error': 'Нет доступа'},
                status=status.HTTP_403_FORBIDDEN
            )

        is_online = request.data.get('is_online')
        if is_online is None:
            return Response(
                {'error': 'Требуется is_online'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.utils import timezone
        from accounts.models import DriverStatus
        
        driver.is_online = bool(is_online)
        
        # Устанавливаем статус водителя в зависимости от is_online
        if driver.is_online:
            # Если водитель становится онлайн и не занят, устанавливаем ONLINE_IDLE
            if not hasattr(driver, 'status') or driver.status in [
                DriverStatus.OFFLINE, 
                None
            ] or (driver.status == DriverStatus.ON_TRIP and not driver.orders.filter(
                status__in=['assigned', 'driver_en_route', 'arrived_waiting', 'ride_ongoing']
            ).exists()):
                driver.status = DriverStatus.ONLINE_IDLE
                driver.idle_since = timezone.now()
        else:
            # Если водитель становится офлайн
            driver.status = DriverStatus.OFFLINE
            driver.idle_since = None
        
        driver.save(update_fields=['is_online', 'status', 'idle_since'])

        return Response(DriverSerializer(driver).data)

    def create(self, request, *args, **kwargs):
        """Создание водителя с возвратом полных данных включая регион"""
        # Используем DriverCreateSerializer для валидации
        create_serializer = DriverCreateSerializer(data=request.data)
        create_serializer.is_valid(raise_exception=True)
        
        validated_data = create_serializer.validated_data
        
        # Проверяем права доступа (только админы могут создавать)
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут создавать водителей')
        
        # Получаем регион
        region_id = validated_data['region_id']
        try:
            region = Region.objects.get(id=region_id)
        except Region.DoesNotExist:
            raise ValidationError({'region_id': 'Регион не найден'})
        
        # Создаем User
        phone = validated_data['phone']
        email = validated_data.get('email', '')
        password = validated_data['password']
        
        username = f'driver_{phone.replace("+", "").replace(" ", "").replace("-", "")}'
        user = User.objects.create_user(
            username=username,
            phone=phone,
            email=email,
            password=password,
            role='driver'
        )
        
        # Создаем Driver
        driver = Driver.objects.create(
            user=user,
            name=validated_data['name'],
            region=region,
            car_model=validated_data['car_model'],
            plate_number=validated_data['plate_number'],
            capacity=validated_data.get('capacity', 4),
            is_online=validated_data.get('is_online', False)
        )
        
        # Перезагружаем водителя с регионом для правильной сериализации
        driver = Driver.objects.select_related('user', 'region').get(id=driver.id)
        
        # Возвращаем ответ с использованием DriverSerializer для включения региона
        response_serializer = DriverSerializer(driver)
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_update(self, serializer):
        """Обновление водителя с обновлением User"""
        driver = serializer.instance
        
        # Проверяем права доступа
        user = self.request.user
        if hasattr(user, 'driver') and driver.id != user.driver.id:
            if not user.is_staff:
                raise PermissionDenied('Нет прав для обновления этого водителя')
        
        validated_data = serializer.validated_data
        
        # Обновляем поля Driver
        if 'name' in validated_data:
            driver.name = validated_data['name']
        if 'region_id' in validated_data:
            try:
                region = Region.objects.get(id=validated_data['region_id'])
                driver.region = region
            except Region.DoesNotExist:
                raise ValidationError({'region_id': 'Регион не найден'})
        if 'car_model' in validated_data:
            driver.car_model = validated_data['car_model']
        if 'plate_number' in validated_data:
            driver.plate_number = validated_data['plate_number']
        if 'capacity' in validated_data:
            driver.capacity = validated_data['capacity']
        if 'is_online' in validated_data:
            driver.is_online = validated_data['is_online']
        
        driver.save()
        
        # Обновляем поля User
        if 'phone' in validated_data:
            phone = validated_data['phone']
            # Проверяем уникальность телефона
            if User.objects.filter(phone=phone).exclude(id=driver.user.id).exists():
                raise ValidationError({'phone': 'Пользователь с таким телефоном уже существует'})
            driver.user.phone = phone
        if 'email' in validated_data:
            driver.user.email = validated_data['email']
        
        driver.user.save()
        
        serializer.instance = driver

    def perform_destroy(self, instance):
        """Удаление водителя (только админы)"""
        # Проверяем права доступа
        if not self.request.user.is_staff:
            raise PermissionDenied('Только администраторы могут удалять водителей')
        
        # Удаление Driver автоматически удалит User (CASCADE)
        instance.delete()

    @action(detail=False, methods=['get'], url_path='template')
    def download_template(self, request):
        """Скачивание шаблона Excel для импорта водителей"""
        try:
            # Проверяем права доступа (только админы могут скачивать шаблон)
            if not request.user.is_staff:
                raise PermissionDenied('Только администраторы могут скачивать шаблон')
            
            # Генерируем шаблон в памяти
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Водители'
            
            # Определяем заголовки колонок
            headers = [
                'Имя',
                'Телефон',
                'Email',
                'Пароль',
                'Регион',
                'Машина',
                'Гос. номер',
                'Вместимость',
                'Водитель онлайн',
                'Рейтинг'
            ]
            
            # Записываем заголовки
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем пример данных
            first_region = Region.objects.first()
            region_title = first_region.title if first_region else 'Алматы'
            
            example_row = [
                'Иванов Иван Иванович',
                '+7 777 123 4567',
                'driver1@invotaxi.kz',
                'password123',
                region_title,
                'Toyota Camry',
                'A 123 BC 02',
                '4',
                'Нет',
                '5.0'
            ]
            
            for col_idx, value in enumerate(example_row, start=1):
                sheet.cell(row=2, column=col_idx).value = value
            
            # Второй пример
            example_row2 = [
                'Петров Петр Петрович',
                '+7 777 765 4321',
                'driver2@invotaxi.kz',
                'securepass123',
                region_title,
                'Hyundai Sonata',
                'B 456 DE 05',
                '4',
                'Да',
                '4.5'
            ]
            
            for col_idx, value in enumerate(example_row2, start=1):
                sheet.cell(row=3, column=col_idx).value = value
            
            # Настраиваем ширину колонок
            column_widths = {
                'A': 25, 'B': 18, 'C': 25, 'D': 20, 'E': 20,
                'F': 18, 'G': 15, 'H': 12, 'I': 18, 'J': 10
            }
            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width
            
            # Добавляем инструкцию на второй лист
            sheet2 = workbook.create_sheet(title='Инструкция')
            instructions = [
                ['ПОЛЕ', 'ОБЯЗАТЕЛЬНО', 'ОПИСАНИЕ', 'ПРИМЕР'],
                ['Имя', 'Да', 'Полное имя водителя', 'Иванов Иван Иванович'],
                ['Телефон', 'Да', 'Телефон в формате +7...', '+7 777 123 4567'],
                ['Email', 'Нет', 'Email адрес (опционально)', 'driver@invotaxi.kz'],
                ['Пароль', 'Да', 'Минимум 8 символов', 'password123'],
                ['Регион', 'Да', 'Название региона или ID региона', 'Алматы'],
                ['Машина', 'Да', 'Модель автомобиля', 'Toyota Camry'],
                ['Гос. номер', 'Да', 'Государственный номер', 'A 123 BC 02'],
                ['Вместимость', 'Да', 'Количество мест (1-20)', '4'],
                ['Водитель онлайн', 'Нет', 'Да/Нет или True/False', 'Нет'],
                ['Рейтинг', 'Нет', 'Рейтинг от 0 до 5', '5.0'],
                ['', '', '', ''],
                ['ПРИМЕЧАНИЯ:', '', '', ''],
                ['1. Все обязательные поля должны быть заполнены', '', '', ''],
                ['2. Телефон должен быть уникальным', '', '', ''],
                ['3. Пароль должен содержать минимум 8 символов', '', '', ''],
                ['4. Для региона можно использовать название или ID', '', '', ''],
                ['5. Вместимость должна быть от 1 до 20 мест', '', '', ''],
                ['6. Для "Водитель онлайн" используйте: Да/Нет, True/False, 1/0', '', '', ''],
            ]
            
            for row_idx, row_data in enumerate(instructions, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet2.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=12, color='FFFFFF')
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif row_idx == 12 and col_idx == 1:
                        cell.font = Font(bold=True, size=11)
            
            sheet2.column_dimensions['A'].width = 20
            sheet2.column_dimensions['B'].width = 12
            sheet2.column_dimensions['C'].width = 50
            sheet2.column_dimensions['D'].width = 25
            
            # Сохраняем в BytesIO
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            
            # Возвращаем файл как HTTP response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="шаблон_импорт_водителей.xlsx"'
            return response
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Ошибка генерации шаблона: {e}', exc_info=True)
            return Response(
                {'error': f'Ошибка генерации шаблона: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='import')
    def import_drivers(self, request):
        """Импорт водителей из Excel файла"""
        # Проверяем права доступа (только админы могут импортировать)
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут импортировать водителей')
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Файл не предоставлен'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Файл должен быть Excel файлом (.xlsx или .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Получаем опции из запроса
        skip_errors = request.data.get('skip_errors', 'false').lower() == 'true'
        dry_run = request.data.get('dry_run', 'false').lower() == 'true'
        
        # Используем команду импорта для обработки файла
        from accounts.management.commands.import_drivers_from_excel import Command as ImportCommand
        from io import BytesIO
        import tempfile
        import os
        
        # Сохраняем файл во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file_path = tmp_file.name
        
        try:
            # Создаем экземпляр команды
            command = ImportCommand()
            
            # Перенаправляем вывод команды
            from io import StringIO
            import sys
            
            output_buffer = StringIO()
            original_stdout = sys.stdout
            sys.stdout = output_buffer
            
            try:
                # Вызываем обработку файла
                command.handle(
                    file_path=tmp_file_path,
                    dry_run=dry_run,
                    skip_errors=skip_errors
                )
            finally:
                sys.stdout = original_stdout
            
            output = output_buffer.getvalue()
            
            # Парсим вывод для получения статистики
            success_count = 0
            created_count = 0
            updated_count = 0
            failed_count = 0
            errors = []
            
            # Пытаемся извлечь статистику из вывода
            import re
            success_match = re.search(r'Успешно обработано: (\d+)', output)
            if success_match:
                success_count = int(success_match.group(1))
            
            created_match = re.search(r'Создано: (\d+)', output)
            if created_match:
                created_count = int(created_match.group(1))
            
            updated_match = re.search(r'Обновлено: (\d+)', output)
            if updated_match:
                updated_count = int(updated_match.group(1))
            
            failed_match = re.search(r'Ошибок: (\d+)', output)
            if failed_match:
                failed_count = int(failed_match.group(1))
            
            return Response({
                'success': True,
                'message': 'Импорт завершен',
                'statistics': {
                    'success_count': success_count,
                    'created_count': created_count,
                    'updated_count': updated_count,
                    'failed_count': failed_count,
                },
                'output': output if failed_count > 0 else None,  # Показываем вывод только при ошибках
            })
            
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            return Response(
                {
                    'success': False,
                    'error': str(e),
                    'traceback': error_trace if request.user.is_superuser else None
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Удаляем временный файл
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)


def get_client_ip(request):
    """Получение IP адреса клиента"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_user_activity(user, action_type, description, request=None, performed_by=None):
    """Вспомогательная функция для логирования активности пользователя"""
    ip_address = None
    if request:
        ip_address = get_client_ip(request)
    
    UserActivityLog.objects.create(
        user=user,
        action_type=action_type,
        description=description,
        ip_address=ip_address,
        performed_by=performed_by or (request.user if request and request.user.is_authenticated else None)
    )


class AdminUserViewSet(viewsets.ModelViewSet):
    """ViewSet для управления админ-пользователями"""
    queryset = User.objects.filter(is_staff=True).prefetch_related('groups').all()
    serializer_class = AdminUserSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Фильтрация и поиск пользователей"""
        queryset = super().get_queryset()
        
        # Фильтрация по роли (через группы)
        role_filter = self.request.query_params.get('role')
        if role_filter:
            if role_filter == 'Admin':
                queryset = queryset.filter(groups__name='admin') | queryset.filter(groups__isnull=True, is_staff=True)
            elif role_filter == 'Dispatcher':
                queryset = queryset.filter(groups__name='dispatcher')
            elif role_filter == 'Operator':
                queryset = queryset.filter(groups__name='operator')
        
        # Фильтрация по статусу
        status_filter = self.request.query_params.get('status')
        if status_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_filter == 'blocked':
            queryset = queryset.filter(is_active=False)
        
        # Поиск по имени, email, телефону
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(username__icontains=search) |
                models.Q(email__icontains=search) |
                models.Q(phone__icontains=search) |
                models.Q(first_name__icontains=search) |
                models.Q(last_name__icontains=search)
            )
        
        return queryset.distinct()
    
    def get_serializer_class(self):
        """Выбор сериализатора в зависимости от действия"""
        if self.action == 'create':
            return AdminUserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return AdminUserUpdateSerializer
        return AdminUserSerializer
    
    def list(self, request, *args, **kwargs):
        """Список пользователей с пагинацией"""
        response = super().list(request, *args, **kwargs)
        
        # Логируем просмотр списка
        log_user_activity(
            user=request.user,
            action_type='login',
            description=f'Просмотр списка пользователей',
            request=request
        )
        
        return response
    
    def create(self, request, *args, **kwargs):
        """Создание нового пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут создавать пользователей')
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        
        # Логируем создание
        log_user_activity(
            user=user,
            action_type='create',
            description=f'Создан новый пользователь: {user.username} ({user.email})',
            request=request
        )
        
        response_serializer = AdminUserSerializer(user)
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    def update(self, request, *args, **kwargs):
        """Обновление пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут обновлять пользователей')
        
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        # Сохраняем старые значения для лога
        old_role = instance.groups.first().name if instance.groups.exists() else 'Admin'
        old_status = instance.is_active
        
        user = serializer.save()
        
        # Определяем новую роль
        new_role = user.groups.first().name if user.groups.exists() else 'Admin'
        
        # Логируем изменения
        changes = []
        if old_role != new_role:
            changes.append(f'роль: {old_role} → {new_role}')
        if old_status != user.is_active:
            changes.append(f'статус: {"активен" if user.is_active else "заблокирован"}')
        
        if changes:
            log_user_activity(
                user=user,
                action_type='update',
                description=f'Обновлен пользователь: {", ".join(changes)}',
                request=request
            )
        
        return Response(AdminUserSerializer(user).data)
    
    def destroy(self, request, *args, **kwargs):
        """Удаление пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут удалять пользователей')
        
        instance = self.get_object()
        
        # Нельзя удалить самого себя
        if instance.id == request.user.id:
            return Response(
                {'error': 'Нельзя удалить самого себя'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Логируем удаление
        log_user_activity(
            user=instance,
            action_type='delete',
            description=f'Удален пользователь: {instance.username} ({instance.email})',
            request=request
        )
        
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['post'], url_path='toggle-status')
    def toggle_status(self, request, pk=None):
        """Блокировка/разблокировка пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут изменять статус пользователей')
        
        user = self.get_object()
        
        # Нельзя заблокировать самого себя
        if user.id == request.user.id:
            return Response(
                {'error': 'Нельзя заблокировать самого себя'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        old_status = user.is_active
        user.is_active = not user.is_active
        user.save()
        
        action_type = 'unblock' if user.is_active else 'block'
        log_user_activity(
            user=user,
            action_type=action_type,
            description=f'Пользователь {"разблокирован" if user.is_active else "заблокирован"}',
            request=request
        )
        
        return Response(AdminUserSerializer(user).data)
    
    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        """Сброс пароля пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут сбрасывать пароли')
        
        user = self.get_object()
        serializer = PasswordResetSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        password = serializer.validated_data['password']
        user.set_password(password)
        user.save()
        
        log_user_activity(
            user=user,
            action_type='password_reset',
            description='Пароль пользователя сброшен',
            request=request
        )
        
        return Response({'message': 'Пароль успешно изменен'}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['post'], url_path='bulk-action')
    def bulk_action(self, request):
        """Массовые операции над пользователями"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут выполнять массовые операции')
        
        serializer = BulkActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user_ids = serializer.validated_data['user_ids']
        action = serializer.validated_data['action']
        
        # Нельзя выполнять действия над самим собой
        if request.user.id in user_ids:
            return Response(
                {'error': 'Нельзя выполнять действие над самим собой'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        users = User.objects.filter(id__in=user_ids, is_staff=True)
        
        if action == 'block':
            users.update(is_active=False)
            for user in users:
                log_user_activity(
                    user=user,
                    action_type='block',
                    description='Пользователь заблокирован (массовая операция)',
                    request=request
                )
            return Response({'message': f'{users.count()} пользователей заблокировано'})
        
        elif action == 'unblock':
            users.update(is_active=True)
            for user in users:
                log_user_activity(
                    user=user,
                    action_type='unblock',
                    description='Пользователь разблокирован (массовая операция)',
                    request=request
                )
            return Response({'message': f'{users.count()} пользователей разблокировано'})
        
        elif action == 'delete':
            count = users.count()
            for user in users:
                log_user_activity(
                    user=user,
                    action_type='delete',
                    description='Пользователь удален (массовая операция)',
                    request=request
                )
            users.delete()
            return Response({'message': f'{count} пользователей удалено'})
    
    @action(detail=True, methods=['get'], url_path='activity-log')
    def activity_log(self, request, pk=None):
        """История активности пользователя"""
        if not request.user.is_staff:
            raise PermissionDenied('Только администраторы могут просматривать историю активности')
        
        user = self.get_object()
        logs = UserActivityLog.objects.filter(user=user).order_by('-created_at')[:50]
        
        serializer = UserActivityLogSerializer(logs, many=True)
        return Response(serializer.data)

