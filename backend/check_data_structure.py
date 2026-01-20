"""Проверка структуры файла data.xlsx"""
import openpyxl

wb = openpyxl.load_workbook('data.xlsx', data_only=True)
print('Листы в файле:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f'\n=== Лист: {sheet_name} ===')
    ws = wb[sheet_name]
    print(f'Всего строк: {ws.max_row}')
    print(f'Всего колонок: {ws.max_column}')
    
    # Первая строка
    first_row = [cell.value for cell in ws[1][:10]]
    print(f'Первая строка (первые 10 колонок): {first_row}')
    
    # Первые 3 строки данных
    if ws.max_row > 1:
        print('\nПервые 3 строки данных:')
        for i in range(2, min(5, ws.max_row + 1)):
            row = [str(cell.value) if cell.value else '' for cell in ws[i][:10]]
            print(f'  Строка {i}: {row[:5]}')  # Показываем только первые 5 колонок
