"""Проверка структуры файла с водителями"""
import openpyxl
from pathlib import Path

file_path = Path(__file__).parent.parent / 'data' / 'водитель (1).xlsx'

if not file_path.exists():
    print(f'Файл не найден: {file_path}')
else:
    print(f'Проверка файла: {file_path}')
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb.active
    
    print(f'\nЛист: {ws.title}')
    print(f'Всего строк: {ws.max_row}')
    print(f'Всего колонок: {ws.max_column}')
    
    # Первая строка (заголовки)
    print('\nЗаголовки (первые 10 колонок):')
    headers = [cell.value for cell in ws[1][:10]]
    for i, h in enumerate(headers, 1):
        if h:
            print(f'  Колонка {i}: {h}')
    
    # Первые 5 строк данных
    print('\nПервые 5 строк данных:')
    for row_num in range(2, min(7, ws.max_row + 1)):
        row = [cell.value for cell in ws[row_num][:10]]
        print(f'\nСтрока {row_num}:')
        for i, val in enumerate(row[:5], 1):
            if val:
                print(f'  Колонка {i}: {val}')
