"""
Скрипт для преобразования файла data.xlsx в формат для импорта водителей

Структура исходного файла (по изображению):
- Номер строки
- Имя водителя
- Телефон
- Регион
- Машина (модель + гос номер вместе)

Формат для импорта:
- Имя
- Телефон
- Email (опционально)
- Пароль
- Регион
- Машина
- Гос. номер
- Вместимость
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys
import re
from pathlib import Path


def parse_car_info(car_string):
    """
    Парсит строку с информацией о машине
    Примеры:
    - "ЭЛАНТРА 745AEN06" -> model="ЭЛАНТРА", plate="745AEN06"
    - "Лада Гранта 175 AFQ/06" -> model="Лада Гранта", plate="175 AFQ/06"
    - "Тойота Камри 2870ВМ" -> model="Тойота Камри", plate="2870ВМ"
    """
    if not car_string or not str(car_string).strip():
        return None, None
    
    car_string = str(car_string).strip()
    
    # Пытаемся найти гос номер в конце (обычно буквы и цифры, возможно с "/")
    # Паттерн: пробел, затем буквы/цифры, возможно "/", затем еще буквы/цифры
    plate_patterns = [
        r'\s+(\d+[А-ЯA-Z]{1,3}\d{2,3})$',  # Цифры + буквы + цифры в конце
        r'\s+(\d+\s*[А-ЯA-Z]{1,3}\s*/\s*\d{2,3})$',  # Формат "123 ABC/06"
        r'\s+(\d{3,4}[А-ЯA-Z]{2,3})$',  # Только цифры и буквы
        r'\s+([А-ЯA-Z]{1,3}\d{3,4})$',  # Буквы в начале
    ]
    
    plate = None
    model = car_string
    
    for pattern in plate_patterns:
        match = re.search(pattern, car_string, re.IGNORECASE)
        if match:
            plate = match.group(1).strip()
            model = car_string[:match.start()].strip()
            break
    
    # Если не нашли паттерн, пытаемся разделить по последнему пробелу
    if not plate:
        parts = car_string.rsplit(' ', 1)
        if len(parts) == 2:
            # Проверяем, похож ли последний элемент на гос номер
            last_part = parts[1]
            if re.match(r'^[\dА-ЯA-Z/]+$', last_part) and len(last_part) >= 3:
                plate = last_part
                model = parts[0]
    
    # Если все еще не нашли, пытаемся найти любую комбинацию букв и цифр в конце
    if not plate:
        match = re.search(r'(\d+[А-ЯA-Z]{1,3}|\d+[А-ЯA-Z]{1,3}/\d{2,3})$', car_string)
        if match:
            plate = match.group(1).strip()
            model = car_string[:match.start()].strip()
    
    # Если ничего не нашли, используем всю строку как модель
    if not plate:
        # Пытаемся найти хотя бы что-то похожее на номер в конце
        words = car_string.split()
        if len(words) > 1:
            last_word = words[-1]
            if re.match(r'^[\dА-ЯA-Z/]+$', last_word):
                plate = last_word
                model = ' '.join(words[:-1])
            else:
                model = car_string
        else:
            model = car_string
    
    # Очищаем пробелы и приводим к верхнему регистру номер
    if plate:
        plate = plate.upper().replace(' ', '')
    
    return model.strip(), plate


def convert_driver_file(input_file, output_file):
    """Конвертирует файл в формат для импорта"""
    print(f'Чтение файла: {input_file}')
    
    # Загружаем исходный файл
    wb_in = openpyxl.load_workbook(input_file, data_only=True)
    ws_in = wb_in.active
    
    # Создаем новый файл
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Водители'
    
    # Заголовки для импорта
    headers = [
        'Имя',
        'Телефон',
        'Email',
        'Пароль',
        'Регион',
        'Машина',
        'Гос. номер',
        'Вместимость',
        'Водитель онлайн',
        'Рейтинг'
    ]
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Определяем колонки в исходном файле (первая строка может быть заголовком или данными)
    first_row = [cell.value for cell in ws_in[1]]
    print(f'Первая строка исходного файла: {first_row[:5]}')
    
    # Определяем, есть ли заголовки или первая строка уже данные
    has_headers = any(str(val).lower() in ['имя', 'name', 'телефон', 'phone'] for val in first_row[:5] if val)
    
    start_row = 2 if has_headers else 1
    
    # Обрабатываем строки
    row_out = 2
    skipped = 0
    
    for row_num in range(start_row, ws_in.max_row + 1):
        row_values = [cell.value for cell in ws_in[row_num]]
        
        # Пропускаем пустые строки
        if not any(val for val in row_values[:5]):
            skipped += 1
            continue
        
        # Определяем колонки (может быть разная структура)
        # Вариант 1: Номер, Имя, Телефон, Регион, Машина
        # Вариант 2: Имя, Телефон, Регион, Машина (без номера)
        
        name = None
        phone = None
        region = None
        car = None
        
        # Пробуем разные варианты структуры
        if len(row_values) >= 5:
            # Вариант с номером строки
            if isinstance(row_values[0], (int, float)) and str(row_values[1]).strip():
                name = str(row_values[1]).strip() if row_values[1] else None
                phone = str(row_values[2]).strip() if row_values[2] else None
                region = str(row_values[3]).strip() if row_values[3] else None
                car = str(row_values[4]).strip() if row_values[4] else None
            # Вариант без номера строки
            elif str(row_values[0]).strip() and not str(row_values[0]).isdigit():
                name = str(row_values[0]).strip() if row_values[0] else None
                phone = str(row_values[1]).strip() if row_values[1] else None
                region = str(row_values[2]).strip() if row_values[2] else None
                car = str(row_values[3]).strip() if row_values[3] else None
        
        # Валидация обязательных полей
        if not name or not phone or not region or not car:
            print(f'  Пропущена строка {row_num}: не все обязательные поля заполнены')
            skipped += 1
            continue
        
        # Очистка телефона (убираем пробелы, оставляем только цифры и +)
        phone_clean = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if not phone_clean.startswith('+'):
            # Если нет +, добавляем +7 для казахстанских номеров
            if phone_clean.startswith('8'):
                phone_clean = '+7' + phone_clean[1:]
            elif phone_clean.startswith('7'):
                phone_clean = '+' + phone_clean
            else:
                phone_clean = '+7' + phone_clean
        
        # Парсим информацию о машине
        car_model, plate_number = parse_car_info(car)
        
        if not car_model or not plate_number:
            print(f'  Предупреждение: строка {row_num}: не удалось распарсить машину "{car}"')
            # Используем всю строку как модель, номер оставляем пустым
            car_model = car
            plate_number = ''
        
        # Генерируем пароль (по умолчанию)
        password = phone_clean.replace('+', '').replace(' ', '')[:8]  # Первые 8 цифр телефона
        if len(password) < 8:
            password = password + '12345678'[:8-len(password)]  # Дополняем до 8 символов
        
        # Записываем данные
        ws_out.cell(row=row_out, column=1).value = name
        ws_out.cell(row=row_out, column=2).value = phone_clean
        ws_out.cell(row=row_out, column=3).value = ''  # Email - пусто
        ws_out.cell(row=row_out, column=4).value = password
        ws_out.cell(row=row_out, column=5).value = region
        ws_out.cell(row=row_out, column=6).value = car_model
        ws_out.cell(row=row_out, column=7).value = plate_number
        ws_out.cell(row=row_out, column=8).value = 4  # Вместимость по умолчанию
        ws_out.cell(row=row_out, column=9).value = 'Нет'  # Онлайн по умолчанию
        ws_out.cell(row=row_out, column=10).value = 5.0  # Рейтинг по умолчанию
        
        row_out += 1
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 25,  # Имя
        'B': 18,  # Телефон
        'C': 25,  # Email
        'D': 20,  # Пароль
        'E': 20,  # Регион
        'F': 18,  # Машина
        'G': 15,  # Гос. номер
        'H': 12,  # Вместимость
        'I': 18,  # Водитель онлайн
        'J': 10   # Рейтинг
    }
    
    for col_letter, width in column_widths.items():
        ws_out.column_dimensions[col_letter].width = width
    
    # Сохраняем файл
    wb_out.save(output_file)
    
    converted = row_out - 2
    print(f'\n✓ Конвертировано строк: {converted}')
    print(f'  Пропущено строк: {skipped}')
    print(f'  Результат сохранен в: {output_file}')
    
    return converted


if __name__ == '__main__':
    # Пути к файлам
    backend_dir = Path(__file__).parent
    input_file = backend_dir / 'data.xlsx'
    output_file = backend_dir / 'data' / 'drivers_for_import.xlsx'
    
    # Создаем директорию data, если её нет
    output_file.parent.mkdir(exist_ok=True)
    
    if not input_file.exists():
        print(f'Ошибка: файл {input_file} не найден!')
        sys.exit(1)
    
    try:
        convert_driver_file(str(input_file), str(output_file))
        print(f'\n✓ Готово! Теперь можно импортировать водителей командой:')
        print(f'  python manage.py import_drivers_from_excel "data/drivers_for_import.xlsx" --skip-errors')
    except Exception as e:
        print(f'\n✗ Ошибка: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)
