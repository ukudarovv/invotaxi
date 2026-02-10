"""
Преобразование файла водителей в формат для импорта

Исходный формат:
- Колонка 1: Номер строки
- Колонка 2: Имя водителя
- Колонка 3: Телефон
- Колонка 4: Регион
- Колонка 5: Машина (модель + гос номер)

Целевой формат:
- Имя, Телефон, Пароль, Регион, Машина, Гос. номер, Вместимость
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from pathlib import Path


def parse_car_info(car_string):
    """Парсит строку с информацией о машине"""
    if not car_string or not str(car_string).strip():
        return None, None
    
    car_string = str(car_string).strip()
    
    # Паттерны для поиска гос номера
    plate_patterns = [
        r'\s+(\d+[А-ЯA-Z]{1,3}\d{2,3})$',
        r'\s+(\d+\s*[А-ЯA-Z]{1,3}\s*/\s*\d{2,3})$',
        r'\s+(\d{3,4}[А-ЯA-Z]{2,3})$',
        r'\s+([А-ЯA-Z]{1,3}\d{3,4})$',
        r'\s+(\d+[А-ЯA-Z]{2,3}\s*\d{2})$',  # Формат "123ABC06"
        r'\s+(\d{3,4}\s*[А-ЯA-Z]{2,3}\s*\d{2})$',  # Формат "745AEN06"
    ]
    
    plate = None
    model = car_string
    
    for pattern in plate_patterns:
        match = re.search(pattern, car_string, re.IGNORECASE)
        if match:
            plate = match.group(1).strip().upper().replace(' ', '')
            model = car_string[:match.start()].strip()
            break
    
    # Если не нашли, пытаемся разделить по последнему пробелу
    if not plate:
        parts = car_string.rsplit(' ', 1)
        if len(parts) == 2:
            last_part = parts[1].strip().upper()
            # Проверяем, похож ли последний элемент на гос номер
            if re.match(r'^[\dА-ЯA-Z/]+$', last_part) and len(last_part) >= 3:
                plate = last_part.replace(' ', '')
                model = parts[0].strip()
    
    # Если все еще не нашли, пробуем найти любую комбинацию букв и цифр в конце
    if not plate:
        match = re.search(r'(\d+[А-ЯA-Z]{1,3}|\d+[А-ЯA-Z]{1,3}/\d{2,3}|[А-ЯA-Z]{1,3}\d{3,4}|\d{3,4}[А-ЯA-Z]{2,3})$', car_string, re.IGNORECASE)
        if match:
            plate = match.group(1).strip().upper().replace(' ', '')
            model = car_string[:match.start()].strip()
    
    if not plate:
        words = car_string.split()
        if len(words) > 1:
            last_word = words[-1].strip().upper()
            if re.match(r'^[\dА-ЯA-Z/]+$', last_word) and len(last_word) >= 3:
                plate = last_word.replace(' ', '')
                model = ' '.join(words[:-1])
            else:
                model = car_string
        else:
            model = car_string
    
    return model.strip(), plate


def convert_file():
    """Преобразует файл в формат для импорта"""
    input_file = Path(__file__).parent.parent / 'data' / 'водитель (1).xlsx'
    output_file = Path(__file__).parent / 'data' / 'drivers_for_import.xlsx'
    
    # Создаем директорию data
    output_file.parent.mkdir(exist_ok=True)
    
    print(f'Чтение файла: {input_file}')
    
    # Загружаем исходный файл
    wb_in = openpyxl.load_workbook(str(input_file), data_only=True)
    ws_in = wb_in.active
    
    # Создаем новый файл
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Водители'
    
    # Заголовки
    headers = [
        'Имя',
        'Телефон',
        'Email',
        'Пароль',
        'Регион',
        'Машина',
        'Гос. номер',
        'Вместимость',
        'Водитель онлайн',
        'Рейтинг'
    ]
    
    # Записываем заголовки
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Обрабатываем строки (начинаем с первой, так как заголовков нет)
    row_out = 2
    converted = 0
    skipped = 0
    
    for row_num in range(1, ws_in.max_row + 1):
        row_values = [cell.value for cell in ws_in[row_num]]
        
        # Пропускаем пустые строки
        if not any(val for val in row_values[:5]):
            skipped += 1
            continue
        
        # Извлекаем данные (колонки: номер, имя, телефон, регион, машина)
        if len(row_values) < 5:
            skipped += 1
            continue
        
        try:
            # Проверяем, что первая колонка - это номер (может быть int или str)
            first_col = str(row_values[0]).strip() if row_values[0] else ''
            if not first_col or not first_col.isdigit():
                # Если первая колонка не номер, возможно, это данные без номера
                name = str(row_values[0]).strip() if row_values[0] else ''
                phone = str(row_values[1]).strip() if row_values[1] else ''
                region = str(row_values[2]).strip() if row_values[2] else ''
                car = str(row_values[3]).strip() if row_values[3] else ''
            else:
                # Стандартный формат: номер, имя, телефон, регион, машина
                name = str(row_values[1]).strip() if row_values[1] else ''
                phone = str(row_values[2]).strip() if row_values[2] else ''
                region = str(row_values[3]).strip() if row_values[3] else ''
                car = str(row_values[4]).strip() if row_values[4] else ''
            
            # Валидация
            if not name or not phone or not region or not car:
                print(f'  Пропущена строка {row_num}: не все обязательные поля заполнены')
                skipped += 1
                continue
            
            # Очистка телефона
            phone_clean = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
            if phone_clean.startswith('8'):
                phone_clean = '+7' + phone_clean[1:]
            elif phone_clean.startswith('7'):
                phone_clean = '+' + phone_clean
            else:
                phone_clean = '+7' + phone_clean
            
            # Парсим машину
            car_model, plate_number = parse_car_info(car)
            
            if not car_model or not plate_number:
                print(f'  Предупреждение: строка {row_num}: не удалось распарсить машину "{car}"')
                car_model = car
                plate_number = ''
            
            # Генерируем пароль (первые 8 цифр телефона)
            password = phone_clean.replace('+', '').replace(' ', '')[:8]
            if len(password) < 8:
                password = password + '12345678'[:8-len(password)]
            
            # Записываем данные
            ws_out.cell(row=row_out, column=1).value = name
            ws_out.cell(row=row_out, column=2).value = phone_clean
            ws_out.cell(row=row_out, column=3).value = ''  # Email
            ws_out.cell(row=row_out, column=4).value = password
            ws_out.cell(row=row_out, column=5).value = region
            ws_out.cell(row=row_out, column=6).value = car_model
            ws_out.cell(row=row_out, column=7).value = plate_number
            ws_out.cell(row=row_out, column=8).value = 4  # Вместимость по умолчанию
            ws_out.cell(row=row_out, column=9).value = 'Нет'  # Онлайн по умолчанию
            ws_out.cell(row=row_out, column=10).value = 5.0  # Рейтинг по умолчанию
            
            row_out += 1
            converted += 1
            
        except Exception as e:
            print(f'  Ошибка в строке {row_num}: {e}')
            skipped += 1
    
    # Настраиваем ширину колонок
    column_widths = {
        'A': 25, 'B': 18, 'C': 25, 'D': 20, 'E': 20,
        'F': 18, 'G': 15, 'H': 12, 'I': 18, 'J': 10
    }
    for col_letter, width in column_widths.items():
        ws_out.column_dimensions[col_letter].width = width
    
    # Сохраняем файл
    wb_out.save(str(output_file))
    
    print(f'\n[OK] Конвертировано строк: {converted}')
    print(f'  Пропущено строк: {skipped}')
    print(f'  Результат сохранен в: {output_file}')
    print(f'\nДля импорта выполните:')
    print(f'  python manage.py import_drivers_from_excel "data/drivers_for_import.xlsx" --skip-errors')


if __name__ == '__main__':
    convert_file()
