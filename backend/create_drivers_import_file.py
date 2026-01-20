"""
Создание файла для импорта водителей на основе данных из изображения

Из изображения видно:
- 63 водителя
- Структура: Номер, Имя, Телефон, Регион, Машина (модель + гос номер)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from pathlib import Path


def parse_car_info(car_string):
    """Парсит строку с информацией о машине"""
    if not car_string or not str(car_string).strip():
        return None, None
    
    car_string = str(car_string).strip()
    
    # Пытаемся найти гос номер в конце (обычно буквы и цифры, возможно с "/")
    plate_patterns = [
        r'\s+(\d+[А-ЯA-Z]{1,3}\d{2,3})$',
        r'\s+(\d+\s*[А-ЯA-Z]{1,3}\s*/\s*\d{2,3})$',
        r'\s+(\d{3,4}[А-ЯA-Z]{2,3})$',
        r'\s+([А-ЯA-Z]{1,3}\d{3,4})$',
    ]
    
    plate = None
    model = car_string
    
    for pattern in plate_patterns:
        match = re.search(pattern, car_string, re.IGNORECASE)
        if match:
            plate = match.group(1).strip()
            model = car_string[:match.start()].strip()
            break
    
    if not plate:
        parts = car_string.rsplit(' ', 1)
        if len(parts) == 2:
            last_part = parts[1]
            if re.match(r'^[\dА-ЯA-Z/]+$', last_part) and len(last_part) >= 3:
                plate = last_part
                model = parts[0]
    
    if not plate:
        match = re.search(r'(\d+[А-ЯA-Z]{1,3}|\d+[А-ЯA-Z]{1,3}/\d{2,3})$', car_string)
        if match:
            plate = match.group(1).strip()
            model = car_string[:match.start()].strip()
    
    if not plate:
        words = car_string.split()
        if len(words) > 1:
            last_word = words[-1]
            if re.match(r'^[\dА-ЯA-Z/]+$', last_word):
                plate = last_word
                model = ' '.join(words[:-1])
            else:
                model = car_string
        else:
            model = car_string
    
    if plate:
        plate = plate.upper().replace(' ', '')
    
    return model.strip(), plate


# Данные водителей из изображения (пример - нужно заполнить все 63)
# Это пример структуры, пользователь должен предоставить полные данные или указать файл
drivers_data = [
    # Формат: [номер, имя, телефон, регион, машина]
    # Пользователь должен указать файл или мы попробуем найти данные в другом месте
]

def create_import_file_from_data(drivers_data, output_file):
    """Создает файл для импорта из данных"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Водители'
    
    # Заголовки
    headers = [
        'Имя',
        'Телефон',
        'Email',
        'Пароль',
        'Регион',
        'Машина',
        'Гос. номер',
        'Вместимость',
        'Водитель онлайн',
        'Рейтинг'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_out = 2
    for driver in drivers_data:
        if len(driver) < 5:
            continue
        
        _, name, phone, region, car = driver[:5]
        
        # Очистка телефона
        phone_clean = str(phone).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if not phone_clean.startswith('+'):
            if phone_clean.startswith('8'):
                phone_clean = '+7' + phone_clean[1:]
            elif phone_clean.startswith('7'):
                phone_clean = '+' + phone_clean
            else:
                phone_clean = '+7' + phone_clean
        
        # Парсим машину
        car_model, plate_number = parse_car_info(str(car))
        
        if not car_model or not plate_number:
            car_model = str(car)
            plate_number = ''
        
        # Пароль
        password = phone_clean.replace('+', '').replace(' ', '')[:8]
        if len(password) < 8:
            password = password + '12345678'[:8-len(password)]
        
        # Записываем
        ws.cell(row=row_out, column=1).value = str(name).strip()
        ws.cell(row=row_out, column=2).value = phone_clean
        ws.cell(row=row_out, column=3).value = ''
        ws.cell(row=row_out, column=4).value = password
        ws.cell(row=row_out, column=5).value = str(region).strip()
        ws.cell(row=row_out, column=6).value = car_model
        ws.cell(row=row_out, column=7).value = plate_number
        ws.cell(row=row_out, column=8).value = 4
        ws.cell(row=row_out, column=9).value = 'Нет'
        ws.cell(row=row_out, column=10).value = 5.0
        
        row_out += 1
    
    # Ширина колонок
    column_widths = {
        'A': 25, 'B': 18, 'C': 25, 'D': 20, 'E': 20,
        'F': 18, 'G': 15, 'H': 12, 'I': 18, 'J': 10
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    wb.save(output_file)
    print(f'✓ Создан файл для импорта: {output_file}')
    print(f'  Количество водителей: {row_out - 2}')


if __name__ == '__main__':
    print("""
    ВНИМАНИЕ: Этот скрипт требует данные о водителях.
    
    У вас есть два варианта:
    1. Использовать существующую команду импорта с файлом в правильном формате
    2. Указать файл с данными о водителях (если он есть в другом месте)
    
    Для создания шаблона используйте:
    python manage.py import_drivers_from_excel --generate-template
    
    Или если у вас есть файл с водителями (даже в другом формате), 
    укажите путь к нему.
    """)
    
    # Проверяем, может быть есть файл с водителями в другом месте
    import os
    possible_files = [
        'data/водитель (1).xlsx',
        '../data/водитель (1).xlsx',
        'водитель.xlsx',
        'drivers.xlsx',
    ]
    
    for f in possible_files:
        if os.path.exists(f):
            print(f'✓ Найден файл: {f}')
            print('Попробуйте запустить команду импорта:')
            print(f'  python manage.py import_drivers_from_excel "{f}" --skip-errors')
            break
