"""Исправление названий регионов в файле импорта"""
import openpyxl
from pathlib import Path
import django
import os
import sys

sys.path.insert(0, str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'invo_backend.settings')
django.setup()

from regions.models import Region

# Загружаем файл
file_path = Path(__file__).parent / 'data' / 'drivers_for_import.xlsx'
wb = openpyxl.load_workbook(str(file_path), data_only=True)
ws = wb.active

# Получаем список регионов из базы
regions_db = {r.title.lower().strip(): r for r in Region.objects.all()}
print(f'Регионов в базе: {len(regions_db)}')
print('\nСписок регионов в базе:')
for title, region in sorted(regions_db.items()):
    print(f'  "{region.title}" (ID: {region.id})')

# Проверяем файл
print('\nПроверка регионов в файле:')
issues = []
row_num = 2

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[4]:  # Регион в колонке 5 (индекс 4)
        row_num += 1
        continue
    
    region_name = str(row[4]).strip()
    region_lower = region_name.lower().strip()
    
    # Пытаемся найти регион
    found = False
    matched_region = None
    
    # Точное совпадение
    if region_lower in regions_db:
        matched_region = regions_db[region_lower]
        found = True
    else:
        # Частичное совпадение
        for db_title, db_region in regions_db.items():
            if region_lower in db_title or db_title in region_lower:
                matched_region = db_region
                found = True
                break
    
    if not found:
        issues.append((row_num, region_name))
        print(f'  Строка {row_num}: "{region_name}" - НЕ НАЙДЕН')
    else:
        # Заменяем на ID региона
        ws.cell(row=row_num, column=5).value = matched_region.id
        print(f'  Строка {row_num}: "{region_name}" -> {matched_region.title} (ID: {matched_region.id})')
    
    row_num += 1

if issues:
    print(f'\n[!] Найдено проблем: {len(issues)}')
    print('Не найдены регионы:')
    for row, name in issues[:10]:
        print(f'  Строка {row}: "{name}"')
else:
    print('\n[OK] Все регионы найдены и обновлены!')

# Сохраняем файл
wb.save(str(file_path))
print(f'\nФайл обновлен: {file_path}')
