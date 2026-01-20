"""
Скрипт для геокодирования адресов с использованием OpenStreetMap Nominatim.

Поддерживает:
- CSV файлы (вход/выход)
- Excel файлы (xlsx) (вход/выход)
- Кэширование результатов (JSON)
- Retry логику при ошибках
- Rate limiting (1 запрос/сек)

Использование:
    python geocode_nominatim.py input.csv output.csv
    python geocode_nominatim.py input.xlsx output.xlsx
    python geocode_nominatim.py  # использует input.csv/output.csv по умолчанию
"""
import csv
import json
import time
import sys
from pathlib import Path
from typing import Dict, Optional, Any

try:
    import requests
except ImportError:
    print("ERROR: requests library not found. Install it with: pip install requests")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("WARNING: openpyxl not installed. Excel support disabled. Install with: pip install openpyxl")

BASE_URL = "https://nominatim.openstreetmap.org/search"
DEFAULT_CACHE_FILE = "cache_nominatim.json"
DEFAULT_INPUT_FILE = "input.csv"
DEFAULT_OUTPUT_FILE = "output.csv"


def normalize_address(address: str) -> str:
    """
    Нормализация адреса для лучшей точности геокодирования.
    Заменяет сокращения на полные формы.
    """
    if not address:
        return ""
    
    address = address.strip()
    # Замены для улучшения качества поиска
    replacements = {
        "ул.": "улица",
        "пр.": "проспект",
        "пр-т": "проспект",
        "пр-кт": "проспект",
        "бул.": "бульвар",
        "б-р": "бульвар",
        "пер.": "переулок",
        "пл.": "площадь",
        "ш.": "шоссе",
        "мкр.": "микрорайон",
        "мкрн": "микрорайон",
    }
    
    for abbrev, full in replacements.items():
        # Заменяем только целые слова
        import re
        pattern = r'\b' + re.escape(abbrev) + r'\b'
        address = re.sub(pattern, full, address, flags=re.IGNORECASE)
    
    # Убираем лишние пробелы и запятые
    address = " ".join(address.split())
    address = address.replace(" ,", ",").replace(", ,", ",").strip()
    
    return address


def geocode_nominatim(
    address: str,
    session: requests.Session,
    countrycodes: str = "kz",
    retry_count: int = 5
) -> Dict[str, Any]:
    """
    Геокодирование одного адреса через Nominatim API.
    
    Args:
        address: Адрес для геокодирования
        session: requests.Session объект
        countrycodes: Коды стран (по умолчанию "kz" - Казахстан)
        retry_count: Количество попыток при ошибках
    
    Returns:
        Словарь с результатами: status, lat, lon, display_name
    """
    normalized_address = normalize_address(address)
    
    params = {
        "q": normalized_address,
        "format": "jsonv2",
        "limit": 1,
        "addressdetails": 1,
        "countrycodes": countrycodes,
    }
    
    for attempt in range(1, retry_count + 1):
        try:
            r = session.get(BASE_URL, params=params, timeout=20)
            
            # Nominatim может отвечать 429 (слишком много запросов)
            if r.status_code == 429:
                wait_time = 2 * attempt
                print(f"  Rate limit (429). Waiting {wait_time}s before retry {attempt}/{retry_count}...")
                time.sleep(wait_time)
                continue
            
            r.raise_for_status()
            data = r.json()
            
            if not data or len(data) == 0:
                return {
                    "status": "not_found",
                    "lat": None,
                    "lon": None,
                    "display_name": None,
                    "original_address": address,
                    "normalized_address": normalized_address
                }
            
            return {
                "status": "ok",
                "lat": data[0].get("lat"),
                "lon": data[0].get("lon"),
                "display_name": data[0].get("display_name"),
                "original_address": address,
                "normalized_address": normalized_address
            }
            
        except requests.RequestException as e:
            if attempt < retry_count:
                wait_time = 2 * attempt
                print(f"  Request error: {e}. Waiting {wait_time}s before retry {attempt}/{retry_count}...")
                time.sleep(wait_time)
            else:
                return {
                    "status": "error",
                    "lat": None,
                    "lon": None,
                    "display_name": None,
                    "error": str(e),
                    "original_address": address,
                    "normalized_address": normalized_address
                }
    
    return {
        "status": "error",
        "lat": None,
        "lon": None,
        "display_name": None,
        "original_address": address,
        "normalized_address": normalized_address
    }


def load_cache(cache_path: Path) -> Dict[str, Dict[str, Any]]:
    """Загрузка кэша из JSON файла."""
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, IOError) as e:
            print(f"WARNING: Could not load cache file {cache_path}: {e}")
            return {}
    return {}


def save_cache(cache_path: Path, cache: Dict[str, Dict[str, Any]]):
    """Сохранение кэша в JSON файл."""
    try:
        cache_path.write_text(
            json.dumps(cache, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except IOError as e:
        print(f"WARNING: Could not save cache file {cache_path}: {e}")


def is_excel_file(filepath: str) -> bool:
    """Проверка, является ли файл Excel файлом."""
    return Path(filepath).suffix.lower() in ['.xlsx', '.xls']


def process_csv(input_file: str, output_file: str, cache_file: str, address_column: Optional[str] = None):
    """Обработка CSV файла."""
    cache_path = Path(cache_file)
    cache = load_cache(cache_path)
    
    # Настройка requests session с User-Agent
    session = requests.Session()
    session.headers.update({
        "User-Agent": "InvoTaxi-GeoCoder/1.0 (contact: admin@invotaxi.kz)"
    })
    
    # Читаем входной CSV
    try:
        with open(input_file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    except FileNotFoundError:
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not read CSV file {input_file}: {e}")
        sys.exit(1)
    
    if not rows:
        print("WARNING: Input file is empty.")
        return
    
    # Определяем колонку с адресами
    if address_column:
        if address_column not in rows[0]:
            print(f"ERROR: Column '{address_column}' not found in CSV.")
            print(f"Available columns: {', '.join(rows[0].keys())}")
            sys.exit(1)
    else:
        # Автоматически ищем колонку с адресом
        possible_columns = ['address', 'адрес', 'address_full', 'full_address', 'location', 'адрес_полный']
        address_column = None
        for col in possible_columns:
            if col in rows[0]:
                address_column = col
                break
        
        if not address_column:
            # Используем первую колонку
            address_column = list(rows[0].keys())[0]
            print(f"WARNING: No 'address' column found. Using first column: '{address_column}'")
    
    # Подготавливаем выходные колонки
    out_fields = list(rows[0].keys())
    for col in ["lat", "lon", "status", "matched_name", "normalized_address"]:
        if col not in out_fields:
            out_fields.append(col)
    
    # Обработка каждой строки
    total = len(rows)
    print(f"\nProcessing {total} addresses from {input_file}...")
    print(f"Using address column: '{address_column}'")
    print(f"Cache file: {cache_file}\n")
    
    for i, row in enumerate(rows, start=1):
        address = (row.get(address_column) or "").strip()
        
        if not address:
            row["status"] = "empty_address"
            row["lat"] = ""
            row["lon"] = ""
            row["matched_name"] = ""
            row["normalized_address"] = ""
            continue
        
        # Проверяем кэш
        cache_key = address.lower().strip()
        if cache_key in cache:
            res = cache[cache_key]
            print(f"[{i}/{total}] ✓ Cached: {address[:60]}...")
        else:
            print(f"[{i}/{total}] → Geocoding: {address[:60]}...", end=" ", flush=True)
            res = geocode_nominatim(address, session=session, countrycodes="kz")
            
            cache[cache_key] = res
            save_cache(cache_path, cache)
            
            status_emoji = "✓" if res["status"] == "ok" else "✗"
            print(f"{status_emoji} {res['status']}")
            
            # Соблюдаем rate limit: 1 запрос/сек
            time.sleep(1.0)
        
        # Заполняем строку результатами
        row["status"] = res.get("status", "")
        row["lat"] = res.get("lat") or ""
        row["lon"] = res.get("lon") or ""
        row["matched_name"] = res.get("display_name") or ""
        row["normalized_address"] = res.get("normalized_address", "")
        
        # Периодически сохраняем кэш (каждые 20 строк)
        if i % 20 == 0:
            save_cache(cache_path, cache)
    
    # Сохраняем финальный кэш
    save_cache(cache_path, cache)
    
    # Пишем выходной CSV
    try:
        with open(output_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=out_fields)
            writer.writeheader()
            writer.writerows(rows)
        print(f"\n✓ Done! Results saved to: {output_file}")
        print(f"✓ Cache saved to: {cache_file}")
    except Exception as e:
        print(f"\nERROR: Could not write output file {output_file}: {e}")
        sys.exit(1)


def process_excel(input_file: str, output_file: str, cache_file: str, address_column: Optional[str] = None, sheet_name: Optional[str] = None):
    """Обработка Excel файла."""
    if not EXCEL_SUPPORT:
        print("ERROR: Excel support is not available. Install openpyxl: pip install openpyxl")
        sys.exit(1)
    
    cache_path = Path(cache_file)
    cache = load_cache(cache_path)
    
    # Настройка requests session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "InvoTaxi-GeoCoder/1.0 (contact: admin@invotaxi.kz)"
    })
    
    # Загружаем Excel файл
    try:
        wb = load_workbook(input_file, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                print(f"ERROR: Sheet '{sheet_name}' not found in Excel file.")
                print(f"Available sheets: {', '.join(wb.sheetnames)}")
                sys.exit(1)
            ws = wb[sheet_name]
        else:
            ws = wb.active
    except FileNotFoundError:
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not read Excel file {input_file}: {e}")
        sys.exit(1)
    
    # Читаем заголовки
    headers = []
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell for cell in row if cell):
            headers = [str(cell).strip() if cell else f"Column_{i+1}" for i, cell in enumerate(row)]
            header_row = row_idx
            break
    
    if not headers:
        print("ERROR: Could not find headers in Excel file.")
        sys.exit(1)
    
    # Определяем колонку с адресами
    if address_column:
        if address_column not in headers:
            print(f"ERROR: Column '{address_column}' not found in Excel.")
            print(f"Available columns: {', '.join(headers)}")
            sys.exit(1)
        address_col_idx = headers.index(address_column)
    else:
        # Автоматически ищем колонку с адресом
        possible_columns = ['address', 'адрес', 'address_full', 'full_address', 'location', 'адрес_полный']
        address_col_idx = None
        for col in possible_columns:
            if col in headers:
                address_col_idx = headers.index(col)
                address_column = col
                break
        
        if address_col_idx is None:
            address_col_idx = 0
            address_column = headers[0]
            print(f"WARNING: No 'address' column found. Using first column: '{address_column}'")
    
    # Читаем данные
    rows_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        row_dict = {}
        for col_idx, cell_value in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = str(cell_value).strip() if cell_value else ""
        rows_data.append((row_idx, row_dict))
    
    if not rows_data:
        print("WARNING: No data rows found in Excel file.")
        return
    
    # Добавляем выходные колонки, если их нет
    for col in ["lat", "lon", "status", "matched_name", "normalized_address"]:
        if col not in headers:
            headers.append(col)
    
    # Обработка каждой строки
    total = len(rows_data)
    print(f"\nProcessing {total} addresses from {input_file} (sheet: {ws.title})...")
    print(f"Using address column: '{address_column}'")
    print(f"Cache file: {cache_file}\n")
    
    for idx, (row_idx, row) in enumerate(rows_data, start=1):
        address = (row.get(address_column) or "").strip()
        
        if not address:
            row["status"] = "empty_address"
            row["lat"] = ""
            row["lon"] = ""
            row["matched_name"] = ""
            row["normalized_address"] = ""
            continue
        
        # Проверяем кэш
        cache_key = address.lower().strip()
        if cache_key in cache:
            res = cache[cache_key]
            print(f"[{idx}/{total}] ✓ Cached: {address[:60]}...")
        else:
            print(f"[{idx}/{total}] → Geocoding: {address[:60]}...", end=" ", flush=True)
            res = geocode_nominatim(address, session=session, countrycodes="kz")
            
            cache[cache_key] = res
            save_cache(cache_path, cache)
            
            status_emoji = "✓" if res["status"] == "ok" else "✗"
            print(f"{status_emoji} {res['status']}")
            
            # Соблюдаем rate limit: 1 запрос/сек
            time.sleep(1.0)
        
        # Заполняем строку результатами
        row["status"] = res.get("status", "")
        row["lat"] = res.get("lat") or ""
        row["lon"] = res.get("lon") or ""
        row["matched_name"] = res.get("display_name") or ""
        row["normalized_address"] = res.get("normalized_address", "")
        
        # Периодически сохраняем кэш
        if idx % 20 == 0:
            save_cache(cache_path, cache)
    
    # Сохраняем финальный кэш
    save_cache(cache_path, cache)
    
    # Создаем новый Excel файл с результатами
    try:
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = ws.title
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, start=1):
            ws_out.cell(row=1, column=col_idx, value=header)
        
        # Записываем данные
        for row_idx, (_, row) in enumerate(rows_data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header, "")
                ws_out.cell(row=row_idx, column=col_idx, value=value)
        
        wb_out.save(output_file)
        print(f"\n✓ Done! Results saved to: {output_file}")
        print(f"✓ Cache saved to: {cache_file}")
    except Exception as e:
        print(f"\nERROR: Could not write output file {output_file}: {e}")
        sys.exit(1)


def main():
    """Главная функция."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Geocode addresses using OpenStreetMap Nominatim API",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python geocode_nominatim.py input.csv output.csv
  python geocode_nominatim.py data.xlsx results.xlsx --column "адрес"
  python geocode_nominatim.py input.csv output.csv --cache my_cache.json
        """
    )
    
    parser.add_argument(
        "input_file",
        nargs="?",
        default=DEFAULT_INPUT_FILE,
        help=f"Input file (CSV or Excel). Default: {DEFAULT_INPUT_FILE}"
    )
    
    parser.add_argument(
        "output_file",
        nargs="?",
        default=DEFAULT_OUTPUT_FILE,
        help=f"Output file (CSV or Excel). Default: {DEFAULT_OUTPUT_FILE}"
    )
    
    parser.add_argument(
        "--cache",
        default=DEFAULT_CACHE_FILE,
        help=f"Cache file path. Default: {DEFAULT_CACHE_FILE}"
    )
    
    parser.add_argument(
        "--column",
        help="Name of the address column (auto-detected if not specified)"
    )
    
    parser.add_argument(
        "--sheet",
        help="Excel sheet name (uses active sheet if not specified)"
    )
    
    args = parser.parse_args()
    
    # Проверяем, что входной файл существует
    if not Path(args.input_file).exists():
        print(f"ERROR: Input file not found: {args.input_file}")
        sys.exit(1)
    
    # Определяем формат файлов
    input_is_excel = is_excel_file(args.input_file)
    output_is_excel = is_excel_file(args.output_file)
    
    if input_is_excel != output_is_excel:
        print("WARNING: Input and output file formats should match (both CSV or both Excel)")
    
    # Обрабатываем файл
    if input_is_excel:
        process_excel(args.input_file, args.output_file, args.cache, args.column, args.sheet)
    else:
        process_csv(args.input_file, args.output_file, args.cache, args.column)


if __name__ == "__main__":
    main()
