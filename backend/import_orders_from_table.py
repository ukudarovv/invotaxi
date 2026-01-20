"""
Скрипт для импорта заказов из Excel/CSV таблицы напрямую в Django ORM.
Автоматически создает пассажиров и заказы, группируя по телефону.

Поддерживает:
- CSV файлы
- Excel файлы (xlsx)
- Группировку заказов по пассажиру
- Автоматическое определение полей
- Автоматическое геокодирование адресов
- Автоматическое определение региона

Использование:
    python import_orders_from_table.py input.csv
    python import_orders_from_table.py input.xlsx --sheet "Лист1"
    python import_orders_from_table.py input.csv --dry-run
"""
import os
import sys
import django
from pathlib import Path

# Настройка Django окружения
BASE_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'invo_backend.settings')
django.setup()

# Теперь можно импортировать Django модели и функции
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.contrib.auth.hashers import make_password
from accounts.models import Passenger, User
from regions.models import Region
from regions.services import get_region_by_coordinates
from orders.models import Order, OrderStatus
from orders.geocoding_service import geocode_address
import csv
import argparse
import time as time_module
from typing import Dict, List, Optional
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("WARNING: openpyxl not installed. Excel support disabled. Install with: pip install openpyxl")


def normalize_phone(phone: str) -> str:
    """Нормализация номера телефона"""
    if not phone:
        return ""
    
    phone = str(phone).strip()
    phone = ''.join(c for c in phone if c.isdigit() or c == '+')
    
    if phone.startswith('8'):
        phone = '+7' + phone[1:]
    elif phone.startswith('7'):
        phone = '+' + phone
    
    if phone.startswith('+7') and len(phone) > 2:
        digits = phone[2:].replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if len(digits) >= 10:
            phone = f'+7 {digits[0:3]} {digits[3:6]} {digits[6:8]} {digits[8:10]}'
    
    return phone.strip()


def detect_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    """Автоматическое определение колонок по заголовкам"""
    headers_lower = [h.lower().strip() for h in headers]
    
    column_map = {
        'passenger_phone': None,
        'passenger_name': None,
        'pickup_address': None,
        'dropoff_address': None,
        'time': None,
        'has_companion': None,
    }
    
    # Телефон
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['телефон', 'phone', 'passenger_phone', 'тел']):
            column_map['passenger_phone'] = idx
            break
    
    # Имя
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['имя', 'name', 'passenger_name', 'passenger', 'пассажир', 'фио', 'ф.и.о.']):
            column_map['passenger_name'] = idx
            break
    
    # Откуда
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['откуда', 'pickup', 'from', 'адрес отправления', 'адрес_отправления']):
            column_map['pickup_address'] = idx
            break
    
    # Куда
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['куда', 'dropoff', 'to', 'адрес назначения', 'адрес_назначения']):
            column_map['dropoff_address'] = idx
            break
    
    # Время
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['время', 'time', 'desired_pickup_time', 'время забора']):
            column_map['time'] = idx
            break
    
    # Сопровождение
    for idx, header in enumerate(headers_lower):
        if any(kw in header for kw in ['сопр', 'companion', 'has_companion', 'сопровождение']):
            column_map['has_companion'] = idx
            break
    
    return column_map


def parse_csv_file(file_path: str) -> List[Dict]:
    """Чтение данных из CSV файла"""
    rows = []
    
    with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        
        if not headers:
            print("ERROR: CSV файл пуст или не содержит заголовков")
            sys.exit(1)
        
        column_map = detect_columns(headers)
        
        print(f"\nОпределенные колонки:")
        for field, idx in column_map.items():
            if idx is not None:
                print(f"  {field}: {headers[idx]}")
        
        if column_map['passenger_phone'] is None:
            print("ERROR: Не найдена колонка с телефоном пассажира")
            sys.exit(1)
        if column_map['pickup_address'] is None:
            print("ERROR: Не найдена колонка с адресом отправления")
            sys.exit(1)
        if column_map['dropoff_address'] is None:
            print("ERROR: Не найдена колонка с адресом назначения")
            sys.exit(1)
        
        for row_num, row in enumerate(reader, start=2):
            if not any(cell.strip() for cell in row if cell):
                continue
            
            order_data = {}
            
            if column_map['passenger_phone'] is not None and len(row) > column_map['passenger_phone']:
                order_data['passenger_phone'] = normalize_phone(row[column_map['passenger_phone']])
            
            if column_map['passenger_name'] is not None and len(row) > column_map['passenger_name']:
                order_data['passenger_name'] = row[column_map['passenger_name']].strip()
            
            if column_map['pickup_address'] is not None and len(row) > column_map['pickup_address']:
                order_data['pickup_address'] = row[column_map['pickup_address']].strip()
            
            if column_map['dropoff_address'] is not None and len(row) > column_map['dropoff_address']:
                order_data['dropoff_address'] = row[column_map['dropoff_address']].strip()
            
            if column_map['time'] is not None and len(row) > column_map['time']:
                order_data['time'] = row[column_map['time']].strip()
            
            if column_map['has_companion'] is not None and len(row) > column_map['has_companion']:
                companion_value = row[column_map['has_companion']].strip().lower()
                order_data['has_companion'] = companion_value in ('да', 'yes', 'true', '1', 'сопр', 'сопр.')
            
            if order_data.get('passenger_phone') and order_data.get('pickup_address') and order_data.get('dropoff_address'):
                rows.append(order_data)
    
    print(f"\nПрочитано заказов: {len(rows)}")
    return rows


def parse_excel_file(file_path: str, sheet_name: Optional[str] = None) -> List[Dict]:
    """Чтение данных из Excel файла"""
    if not EXCEL_SUPPORT:
        print("ERROR: Excel support is not available. Install openpyxl: pip install openpyxl")
        sys.exit(1)
    
    rows = []
    
    wb = load_workbook(file_path, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"ERROR: Лист '{sheet_name}' не найден.")
            print(f"Доступные листы: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    # Читаем заголовки
    headers = []
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell for cell in row if cell):
            headers = [str(cell).strip() if cell else "" for cell in row]
            header_row = row_idx
            break
    
    if not headers:
        print("ERROR: Не найдены заголовки в Excel файле")
        sys.exit(1)
    
    column_map = detect_columns(headers)
    
    print(f"\nОпределенные колонки:")
    for field, idx in column_map.items():
        if idx is not None:
            print(f"  {field}: {headers[idx]}")
    
    if column_map['passenger_phone'] is None:
        print("ERROR: Не найдена колонка с телефоном пассажира")
        sys.exit(1)
    if column_map['pickup_address'] is None:
        print("ERROR: Не найдена колонка с адресом отправления")
        sys.exit(1)
    if column_map['dropoff_address'] is None:
        print("ERROR: Не найдена колонка с адресом назначения")
        sys.exit(1)
    
    # Читаем данные
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(cell for cell in row if cell):
            continue
        
        order_data = {}
        
        if column_map['passenger_phone'] is not None and len(row) > column_map['passenger_phone']:
            order_data['passenger_phone'] = normalize_phone(str(row[column_map['passenger_phone']]) if row[column_map['passenger_phone']] else "")
        
        if column_map['passenger_name'] is not None and len(row) > column_map['passenger_name']:
            order_data['passenger_name'] = str(row[column_map['passenger_name']]).strip() if row[column_map['passenger_name']] else ""
        
        if column_map['pickup_address'] is not None and len(row) > column_map['pickup_address']:
            order_data['pickup_address'] = str(row[column_map['pickup_address']]).strip() if row[column_map['pickup_address']] else ""
        
        if column_map['dropoff_address'] is not None and len(row) > column_map['dropoff_address']:
            order_data['dropoff_address'] = str(row[column_map['dropoff_address']]).strip() if row[column_map['dropoff_address']] else ""
        
        if column_map['time'] is not None and len(row) > column_map['time']:
            order_data['time'] = str(row[column_map['time']]).strip() if row[column_map['time']] else ""
        
        if column_map['has_companion'] is not None and len(row) > column_map['has_companion']:
            companion_value = str(row[column_map['has_companion']]).strip().lower() if row[column_map['has_companion']] else ""
            order_data['has_companion'] = companion_value in ('да', 'yes', 'true', '1', 'сопр', 'сопр.')
        
        if order_data.get('passenger_phone') and order_data.get('pickup_address') and order_data.get('dropoff_address'):
            rows.append(order_data)
    
    print(f"\nПрочитано заказов: {len(rows)}")
    return rows


def group_orders_by_passenger(orders: List[Dict]) -> Dict[str, List[Dict]]:
    """Группирует заказы по телефону пассажира"""
    grouped = {}
    
    for order in orders:
        phone = order.get('passenger_phone')
        if phone:
            if phone not in grouped:
                grouped[phone] = []
            grouped[phone].append(order)
    
    return grouped


def get_or_create_passenger(phone: str, first_order: Dict) -> Passenger:
    """Получает или создает пассажира"""
    try:
        user = User.objects.get(phone=phone)
        if hasattr(user, 'passenger'):
            return user.passenger
    except User.DoesNotExist:
        pass
    
    # Пассажира нет, создаем
    # Геокодируем первый заказ для определения региона
    pickup_address = first_order.get('pickup_address')
    pickup_lat = None
    pickup_lon = None
    
    if pickup_address:
        print(f'  Геокодирование адреса для определения региона: {pickup_address}')
        geocode_result = geocode_address(pickup_address)
        if geocode_result['status'] == 'ok':
            pickup_lat = geocode_result['lat']
            pickup_lon = geocode_result['lon']
            time_module.sleep(1.0)  # Rate limit
        else:
            print(f'  Предупреждение: не удалось геокодировать адрес: {geocode_result.get("error", "Адрес не найден")}')
    
    # Определяем регион
    region = None
    if pickup_lat and pickup_lon:
        region = get_region_by_coordinates(pickup_lat, pickup_lon)
        if region:
            print(f'  Регион определен по координатам: {region.id} ({region.title})')
    
    if not region:
        region = Region.objects.first()
        if region:
            print(f'  Использован первый доступный регион: {region.id} ({region.title})')
    
    if not region:
        raise Exception('Не удалось определить регион. Убедитесь, что в БД есть регионы.')
    
    # Создаем пользователя
    username = f'passenger_{phone.replace("+", "").replace(" ", "").replace("-", "").replace("(", "").replace(")", "")}'
    user, _ = User.objects.get_or_create(
        phone=phone,
        defaults={
            'username': username,
            'role': 'passenger',
            'password': make_password(None)
        }
    )
    
    # Создаем пассажира
    passenger_name = first_order.get('passenger_name') or f'Пассажир {phone}'
    disability_category = 'III группа'  # Дефолт
    allowed_companion = first_order.get('has_companion', False)
    
    passenger, created = Passenger.objects.get_or_create(
        user=user,
        defaults={
            'full_name': passenger_name,
            'region': region,
            'disability_category': disability_category,
            'allowed_companion': allowed_companion
        }
    )
    
    if created:
        print(f'  Создан пассажир: {passenger.full_name} (регион: {region.title})')
    
    return passenger


def create_order(passenger: Passenger, order_data: Dict) -> Order:
    """Создает заказ"""
    # Подготовка данных
    pickup_address = order_data.get('pickup_address')
    dropoff_address = order_data.get('dropoff_address')
    
    # Парсим время
    time_str = order_data.get('time', '')
    if time_str:
        try:
            if ':' in time_str and len(time_str) <= 5:
                # Формат "HH:MM"
                time_parts = time_str.split(':')
                hour = int(time_parts[0])
                minute = int(time_parts[1])
                today = date.today()
                desired_pickup_time = timezone.make_aware(datetime.combine(today, datetime.min.time().replace(hour=hour, minute=minute)))
            else:
                desired_pickup_time = parse_datetime(time_str) or timezone.now()
        except:
            desired_pickup_time = timezone.now()
    else:
        desired_pickup_time = timezone.now()
    
    # Геокодируем адреса
    pickup_lat = None
    pickup_lon = None
    dropoff_lat = None
    dropoff_lon = None
    
    # Геокодируем адрес отправления
    if pickup_address:
        geocode_result = geocode_address(pickup_address)
        if geocode_result['status'] == 'ok':
            pickup_lat = geocode_result['lat']
            pickup_lon = geocode_result['lon']
        else:
            raise Exception(f"Не удалось геокодировать адрес отправления: {geocode_result.get('error', 'Адрес не найден')}")
        time_module.sleep(1.0)  # Rate limit
    
    # Геокодируем адрес назначения
    if dropoff_address:
        geocode_result = geocode_address(dropoff_address)
        if geocode_result['status'] == 'ok':
            dropoff_lat = geocode_result['lat']
            dropoff_lon = geocode_result['lon']
        else:
            raise Exception(f"Не удалось геокодировать адрес назначения: {geocode_result.get('error', 'Адрес не найден')}")
    
    # Создаем заказ напрямую через Django ORM
    order_id = f'order_{int(time_module.time() * 1000)}'
    
    order = Order.objects.create(
        id=order_id,
        passenger=passenger,
        pickup_title=pickup_address,
        dropoff_title=dropoff_address,
        pickup_lat=pickup_lat,
        pickup_lon=pickup_lon,
        dropoff_lat=dropoff_lat,
        dropoff_lon=dropoff_lon,
        desired_pickup_time=desired_pickup_time,
        has_companion=order_data.get('has_companion', passenger.allowed_companion),
        note='',
        status=OrderStatus.SUBMITTED
    )
    
    # Рассчитываем предварительную цену
    try:
        from orders.services import PriceCalculator
        price_data = PriceCalculator.calculate_estimated_price(order)
        order.distance_km = price_data['distance_km']
        order.waiting_time_minutes = price_data['waiting_time_minutes']
        order.estimated_price = price_data['estimated_price']
        order.price_breakdown = price_data['price_breakdown']
        order.save()
    except Exception as e:
        # Если расчет цены не удался, оставляем заказ без цены
        print(f'    Предупреждение: не удалось рассчитать цену для заказа {order.id}: {e}')
    
    return order


def main():
    parser = argparse.ArgumentParser(
        description='Импорт заказов из Excel/CSV таблицы напрямую в Django ORM',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Примеры:
  python import_orders_from_table.py data.csv
  python import_orders_from_table.py data.xlsx --sheet "Лист1"
  python import_orders_from_table.py data.csv --dry-run
        """
    )
    
    parser.add_argument(
        'input_file',
        help='Входной файл (CSV или Excel)'
    )
    
    parser.add_argument(
        '--sheet',
        help='Название листа Excel (по умолчанию используется активный лист)'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Только проверка данных без создания заказов'
    )
    
    args = parser.parse_args()
    
    input_file = Path(args.input_file)
    if not input_file.exists():
        print(f"ERROR: Файл не найден: {input_file}")
        sys.exit(1)
    
    print(f"Обработка файла: {input_file}")
    
    # Определяем формат файла
    is_excel = input_file.suffix.lower() in ['.xlsx', '.xls']
    
    # Читаем данные
    if is_excel:
        orders = parse_excel_file(str(input_file), args.sheet)
    else:
        orders = parse_csv_file(str(input_file))
    
    if not orders:
        print("ERROR: Не найдено ни одного заказа в файле")
        sys.exit(1)
    
    # Группируем по пассажиру
    grouped = group_orders_by_passenger(orders)
    print(f"\nНайдено пассажиров: {len(grouped)}")
    print(f"Всего заказов: {len(orders)}")
    
    if args.dry_run:
        print("\n=== DRY RUN MODE ===")
        for phone, passenger_orders in grouped.items():
            print(f'\nПассажир: {phone}')
            print(f'  Имя: {passenger_orders[0].get("passenger_name", "не указано")}')
            print(f'  Заказов: {len(passenger_orders)}')
            for idx, order in enumerate(passenger_orders, 1):
                print(f'    {idx}. {order.get("pickup_address")} -> {order.get("dropoff_address")} ({order.get("time", "без времени")})')
        print("\nИспользуйте без --dry-run для реального создания")
        return
    
    # Создаем заказы
    total_created = 0
    total_errors = 0
    
    for phone, passenger_orders in grouped.items():
        try:
            # Получаем или создаем пассажира
            passenger = get_or_create_passenger(phone, passenger_orders[0])
            
            # Создаем заказы
            created_count = 0
            for idx, order_data in enumerate(passenger_orders, 1):
                try:
                    order = create_order(passenger, order_data)
                    created_count += 1
                    if idx % 10 == 0:
                        print(f'    Обработано {idx}/{len(passenger_orders)} заказов...')
                except Exception as e:
                    error_msg = str(e)
                    if len(error_msg) > 200:
                        error_msg = error_msg[:200] + '...'
                    print(f'    Предупреждение: не удалось создать заказ {idx}: {error_msg}')
            
            total_created += created_count
            print(f'[OK] Пассажир {phone}: создано {created_count} заказов')
            
        except Exception as e:
            total_errors += len(passenger_orders)
            print(f'[ERROR] Ошибка для пассажира {phone}: {e}')
    
    print(f'\n[OK] Успешно создано заказов: {total_created}')
    if total_errors > 0:
        print(f'[ERROR] Ошибок: {total_errors}')


if __name__ == "__main__":
    main()
