"""
Management command для импорта заказов из Excel файла
"""
from django.core.management.base import BaseCommand
from django.utils import timezone
from django.conf import settings
from orders.models import Order, OrderStatus
from orders.serializers import OrderSerializer
from accounts.models import Passenger, User
from regions.services import get_region_by_coordinates
import openpyxl
import logging
from datetime import datetime
import os

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Импортирует заказы из Excel файла (data/data.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'file_path',
            type=str,
            nargs='?',
            default='data/data.xlsx',
            help='Путь к Excel файлу (по умолчанию: data/data.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Только валидация, без создания заказов',
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Пропускать строки с ошибками и продолжать',
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']
        skip_errors = options['skip_errors']
        
        # Разрешаем путь относительно корня проекта (родительская папка backend)
        if not os.path.isabs(file_path):
            # Если путь относительный, разрешаем относительно корня проекта
            project_root = settings.BASE_DIR.parent  # Родитель backend (корень проекта)
            # Преобразуем Path в строку и нормализуем разделители
            file_path = str(project_root / file_path)
        
        # Нормализуем путь (для Windows и Linux)
        file_path = os.path.normpath(file_path)
        
        self.stdout.write(f'Импорт заказов из {file_path}...')
        if dry_run:
            self.stdout.write(self.style.WARNING('Режим валидации (dry-run): заказы не будут созданы'))
        self.stdout.write('=' * 60)
        
        try:
            # Открываем Excel файл
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Определяем заголовки (пробуем первую и вторую строку)
            # Часто в Excel файлах первая строка может быть пустой или содержать метаданные
            headers = []
            header_row = 1
            
            # Пробуем первую строку
            first_row_values = []
            for cell in sheet[1]:
                if cell.value:
                    value = str(cell.value).strip().lower()
                    first_row_values.append(value)
                else:
                    first_row_values.append('')
            
            # Если первая строка пустая, содержит только "1", или очень мало значений - пробуем вторую
            non_empty_first = [v for v in first_row_values if v and v != '1']
            if len(non_empty_first) < 3:
                header_row = 2
                for cell in sheet[2]:
                    if cell.value:
                        value = str(cell.value).strip().lower()
                    else:
                        value = ''
                    headers.append(value)
            else:
                headers = first_row_values
            
            self.stdout.write(f'Найдено колонок: {len(headers)}')
            self.stdout.write(f'Используется строка {header_row} как заголовки')
            self.stdout.write(f'Заголовки: {", ".join(headers[:10])}...' if len(headers) > 10 else f'Заголовки: {", ".join(headers)}')
            self.stdout.write('')
            
            # Маппинг возможных названий колонок (частичные совпадения)
            column_mapping = {
                'passenger_id': ['passenger_id', 'passenger', 'id пассажира', 'id'],
                'passenger_phone': ['passenger_phone', 'phone', 'телефон', 'phone пассажира', 'телефон пассажира'],
                'pickup_title': ['pickup_title', 'адрес отправления', 'откуда', 'pickup address', 'адрес отправ'],
                'pickup_lat': ['pickup_lat', 'широта отправления', 'lat отправления', 'pickup_latitude', 'lat отпр', 'широта отпр'],
                'pickup_lon': ['pickup_lon', 'долгота отправления', 'lon отправления', 'pickup_longitude', 'lon отпр', 'долгота отпр'],
                'dropoff_title': ['dropoff_title', 'адрес назначения', 'куда', 'dropoff address', 'адрес назнач'],
                'dropoff_lat': ['dropoff_lat', 'широта назначения', 'lat назначения', 'dropoff_latitude', 'lat назнач', 'широта назнач'],
                'dropoff_lon': ['dropoff_lon', 'долгота назначения', 'lon назначения', 'dropoff_longitude', 'lon назнач', 'долгота назнач'],
                'desired_pickup_time': ['desired_pickup_time', 'время забора', 'дата и время', 'pickup time', 'время', 'дата'],
                'has_companion': ['has_companion', 'с сопровождением', 'companion', 'сопровожден'],
                'note': ['note', 'примечание', 'комментарий'],
                'status': ['status', 'статус']
            }
            
            # Создаем маппинг индексов колонок (с поддержкой частичных совпадений)
            column_indices = {}
            for key, possible_names in column_mapping.items():
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    # Проверяем точное совпадение
                    if header in possible_names:
                        column_indices[key] = idx
                        break
                    # Проверяем частичное совпадение (содержится в заголовке)
                    for name in possible_names:
                        if name in header or header in name:
                            column_indices[key] = idx
                            break
                    if key in column_indices:
                        break
            
            self.stdout.write('Маппинг колонок:')
            for key, idx in column_indices.items():
                self.stdout.write(f'  {key} -> колонка {idx + 1} ({headers[idx]})')
            self.stdout.write('')
            
            # Проверяем обязательные поля
            # Если координаты отсутствуют, можно использовать адреса (но нужен геокодинг)
            required_fields = ['pickup_title', 'dropoff_title']
            missing_fields = [f for f in required_fields if f not in column_indices]
            if missing_fields:
                self.stdout.write(self.style.ERROR(f'Отсутствуют обязательные колонки: {", ".join(missing_fields)}'))
                return
            
            # Проверяем координаты (предупреждение, но не ошибка)
            coord_fields = ['pickup_lat', 'pickup_lon', 'dropoff_lat', 'dropoff_lon']
            missing_coords = [f for f in coord_fields if f not in column_indices]
            if missing_coords:
                self.stdout.write(self.style.WARNING(
                    f'ВНИМАНИЕ: Отсутствуют колонки с координатами: {", ".join(missing_coords)}. '
                    f'Координаты будут установлены в 0,0. Рекомендуется добавить координаты или использовать геокодинг.'
                ))
            
            # Обрабатываем строки
            success_count = 0
            failed_count = 0
            errors = []
            imported_ids = []
            
            start_row = header_row + 1  # Начинаем со строки после заголовков
            for row_num in range(start_row, sheet.max_row + 1):
                try:
                    row_data = {}
                    for key, idx in column_indices.items():
                        cell_value = sheet.cell(row=row_num, column=idx + 1).value
                        if cell_value is not None:
                            row_data[key] = str(cell_value).strip()
                        else:
                            row_data[key] = ''
                    
                    # Валидация и создание заказа
                    order_data = self._parse_order_row(row_data, row_num, column_indices, headers)
                    
                    if dry_run:
                        success_count += 1
                        continue
                    
                    # Создаем заказ через сериализатор
                    serializer = OrderSerializer(data=order_data)
                    if serializer.is_valid():
                        order = serializer.save()
                        success_count += 1
                        imported_ids.append(order.id)
                        self.stdout.write(self.style.SUCCESS(f'  Строка {row_num}: Заказ {order.id} создан'))
                    else:
                        raise ValueError(f"Валидация не пройдена: {serializer.errors}")
                        
                except Exception as e:
                    failed_count += 1
                    error_msg = f'Строка {row_num}: {str(e)}'
                    errors.append({'row': row_num, 'message': str(e)})
                    self.stdout.write(self.style.ERROR(f'  {error_msg}'))
                    
                    if not skip_errors:
                        self.stdout.write(self.style.ERROR(f'Импорт остановлен из-за ошибки в строке {row_num}'))
                        break
            
            # Итоговый отчет
            self.stdout.write('')
            self.stdout.write('=' * 60)
            self.stdout.write(self.style.SUCCESS(f'Успешно обработано: {success_count}'))
            self.stdout.write(self.style.ERROR(f'Ошибок: {failed_count}'))
            
            if errors and (skip_errors or dry_run):
                self.stdout.write('')
                self.stdout.write('Детали ошибок:')
                for error in errors[:20]:  # Показываем первые 20
                    self.stdout.write(f"  Строка {error['row']}: {error['message']}")
                if len(errors) > 20:
                    self.stdout.write(f'  ... и еще {len(errors) - 20} ошибок')
            
            if imported_ids:
                self.stdout.write('')
                self.stdout.write(f'Созданные заказы: {", ".join(imported_ids[:10])}')
                if len(imported_ids) > 10:
                    self.stdout.write(f'... и еще {len(imported_ids) - 10} заказов')
                    
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f'Файл не найден: {file_path}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Ошибка импорта: {str(e)}'))
            logger.error(f'Ошибка импорта заказов из Excel: {e}', exc_info=True)
    
    def _parse_order_row(self, row_data: dict, row_num: int, column_indices: dict, headers: list) -> dict:
        """Парсит строку Excel и возвращает данные для создания заказа"""
        # Находим пассажира
        passenger_id = row_data.get('passenger_id', '').strip()
        passenger_phone = row_data.get('passenger_phone', '').strip()
        
        if not passenger_id and not passenger_phone:
            raise ValueError('Не указан passenger_id или passenger_phone')
        
        passenger = None
        if passenger_id:
            try:
                passenger = Passenger.objects.get(id=int(passenger_id))
            except (Passenger.DoesNotExist, ValueError):
                raise ValueError(f'Пассажир с ID {passenger_id} не найден')
        elif passenger_phone:
            try:
                user = User.objects.get(phone=passenger_phone)
                if hasattr(user, 'passenger'):
                    passenger = user.passenger
                else:
                    raise ValueError(f'Пользователь с телефоном {passenger_phone} не является пассажиром')
            except User.DoesNotExist:
                raise ValueError(f'Пользователь с телефоном {passenger_phone} не найден')
        
        # Координаты
        try:
            pickup_lat = float(row_data.get('pickup_lat', 0))
            pickup_lon = float(row_data.get('pickup_lon', 0))
            dropoff_lat = float(row_data.get('dropoff_lat', 0))
            dropoff_lon = float(row_data.get('dropoff_lon', 0))
        except (ValueError, TypeError):
            raise ValueError('Неверный формат координат (должны быть числа)')
        
        if not pickup_lat or not pickup_lon or not dropoff_lat or not dropoff_lon:
            raise ValueError('Не указаны координаты pickup или dropoff')
        
        # Адреса
        pickup_title = row_data.get('pickup_title', '').strip() or f'{pickup_lat}, {pickup_lon}'
        dropoff_title = row_data.get('dropoff_title', '').strip() or f'{dropoff_lat}, {dropoff_lon}'
        
        # Время забора
        desired_pickup_time_str = row_data.get('desired_pickup_time', '').strip()
        if desired_pickup_time_str:
            try:
                # Пытаемся распарсить дату из Excel (может быть datetime объектом или строкой)
                if isinstance(desired_pickup_time_str, datetime):
                    desired_pickup_time = timezone.make_aware(desired_pickup_time_str)
                else:
                    from django.utils.dateparse import parse_datetime
                    parsed = parse_datetime(desired_pickup_time_str)
                    if parsed:
                        desired_pickup_time = parsed
                    else:
                        # Пробуем разные форматы
                        try:
                            dt = datetime.strptime(desired_pickup_time_str, '%Y-%m-%d %H:%M:%S')
                            desired_pickup_time = timezone.make_aware(dt)
                        except:
                            dt = datetime.strptime(desired_pickup_time_str, '%Y-%m-%d')
                            desired_pickup_time = timezone.make_aware(dt)
            except Exception as e:
                raise ValueError(f'Неверный формат даты: {desired_pickup_time_str} ({str(e)})')
        else:
            desired_pickup_time = timezone.now()
        
        # Опциональные поля
        has_companion_str = row_data.get('has_companion', '').strip().lower()
        has_companion = has_companion_str in ('true', '1', 'yes', 'да', 'д')
        note = row_data.get('note', '').strip()
        
        # Начальный статус
        status_str = row_data.get('status', '').strip()
        if status_str and status_str in dict(OrderStatus.choices):
            initial_status = status_str
        else:
            initial_status = OrderStatus.CREATED
        
        return {
            'passenger': passenger.id,
            'pickup_title': pickup_title,
            'dropoff_title': dropoff_title,
            'pickup_lat': pickup_lat,
            'pickup_lon': pickup_lon,
            'dropoff_lat': dropoff_lat,
            'dropoff_lon': dropoff_lon,
            'desired_pickup_time': desired_pickup_time.isoformat(),
            'has_companion': has_companion,
            'note': note,
            'status': initial_status
        }
